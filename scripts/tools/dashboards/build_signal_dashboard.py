from __future__ import annotations

import argparse
import json
import math
import re
import sqlite3
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from nptdms import TdmsFile
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
INPUT_ROOT = REPO_ROOT / "data" / "derived" / "small_overlap" / "preprocessed_signals"
OUTPUT_ROOT = REPO_ROOT / "output" / "small_overlap" / "dashboard"
PLOTS_ROOT = REPO_ROOT / "output" / "small_overlap" / "plots"
RESEARCH_DB = REPO_ROOT / "data" / "research" / "research.sqlite"
MAX_PLOT_POINTS = 1200


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a static HTML dashboard for one IIHS signal case.")
    parser.add_argument("--filegroup-id", type=int, required=True)
    parser.add_argument("--test-code", default=None)
    parser.add_argument("--input-root", default=None)
    parser.add_argument("--output-root", default=None)
    parser.add_argument("--plots-root", default=None)
    parser.add_argument("--research-db", default=None)
    return parser.parse_args()


def resolve_repo_path(value: str | None, default: Path) -> Path:
    if not value:
        return default
    path = Path(value)
    return path if path.is_absolute() else REPO_ROOT / path


def find_case_root(input_root: Path, filegroup_id: int, test_code: str | None) -> Path:
    if test_code:
        candidate = input_root / f"{filegroup_id}-{test_code}"
        if candidate.exists():
            return candidate
    matches = sorted(input_root.glob(f"{filegroup_id}-*"))
    if not matches:
        raise FileNotFoundError(f"preprocessed case root not found for filegroup_id={filegroup_id}")
    return matches[0]


def load_manifest(case_root: Path) -> dict[str, Any]:
    return json.loads((case_root / "preprocessing_manifest.json").read_text(encoding="utf-8"))


def downsample(x_values: np.ndarray, y_values: np.ndarray, max_points: int = MAX_PLOT_POINTS) -> tuple[np.ndarray, np.ndarray]:
    if len(y_values) <= max_points:
        return x_values, y_values
    stride = max(1, math.ceil(len(y_values) / max_points))
    return x_values[::stride], y_values[::stride]


def to_json_list(values: np.ndarray, digits: int = 6) -> list[float | None]:
    output: list[float | None] = []
    for value in values:
        if value is None:
            output.append(None)
            continue
        numeric = float(value)
        if not math.isfinite(numeric):
            output.append(None)
            continue
        output.append(round(numeric, digits))
    return output


def summarize_series(x_values: np.ndarray, y_values: np.ndarray, x_label: str) -> dict[str, Any]:
    finite_mask = np.isfinite(y_values)
    finite_y = y_values[finite_mask] if finite_mask.any() else y_values
    if finite_y.size == 0:
        return {
            "sample_count": int(y_values.size),
            "x_label": x_label,
            "min": None,
            "max": None,
            "mean": None,
            "std": None,
            "x_start": None,
            "x_end": None,
        }
    min_idx = int(np.nanargmin(y_values))
    max_idx = int(np.nanargmax(y_values))
    return {
        "sample_count": int(y_values.size),
        "x_label": x_label,
        "min": round(float(y_values[min_idx]), 6),
        "max": round(float(y_values[max_idx]), 6),
        "mean": round(float(np.nanmean(y_values)), 6),
        "std": round(float(np.nanstd(y_values)), 6),
        "x_start": round(float(x_values[0]), 6) if x_values.size else None,
        "x_end": round(float(x_values[-1]), 6) if x_values.size else None,
        "x_at_min": round(float(x_values[min_idx]), 6) if x_values.size > min_idx else None,
        "x_at_max": round(float(x_values[max_idx]), 6) if x_values.size > max_idx else None,
    }


def detect_group_time_axis(tdms: TdmsFile, group_name: str) -> np.ndarray | None:
    group = tdms[group_name]
    channel_names = {channel.name for channel in group.channels()}
    if group_name == "11_Corridors" and "11Time" in channel_names:
        values = np.asarray(group["11Time"][:], dtype=float)
        return values if values.size else None
    if group_name == "CEN2005_Raw_Data" and "Time axis" in channel_names:
        values = np.asarray(group["Time axis"][:], dtype=float)
        return values if values.size else None
    return None


def resolve_channel_x(channel: Any, explicit_group_time: np.ndarray | None) -> tuple[np.ndarray, str]:
    values = channel[:]
    try:
        time_track = np.asarray(channel.time_track(), dtype=float)
        if time_track.size == len(values):
            return time_track, "time_s"
    except Exception:
        pass
    if explicit_group_time is not None and explicit_group_time.size == len(values):
        return explicit_group_time, "time_s"
    return np.arange(len(values), dtype=float), "sample_index"


def build_plot_channel(
    channel_name: str,
    unit: str,
    x_values: np.ndarray,
    y_values: np.ndarray,
    x_label: str,
) -> dict[str, Any]:
    plot_x, plot_y = downsample(x_values, y_values)
    return {
        "name": channel_name,
        "unit": unit,
        "x_label": x_label,
        "x": to_json_list(plot_x),
        "y": to_json_list(plot_y),
        "stats": summarize_series(x_values, y_values, x_label),
        "plotted_points": int(len(plot_x)),
        "full_points": int(len(y_values)),
    }


def build_scalar_channel(channel_name: str, unit: str, values: np.ndarray) -> dict[str, Any]:
    return {
        "name": channel_name,
        "unit": unit,
        "values": to_json_list(values),
        "sample_count": int(values.size),
        "min": round(float(np.nanmin(values)), 6) if values.size else None,
        "max": round(float(np.nanmax(values)), 6) if values.size else None,
        "mean": round(float(np.nanmean(values)), 6) if values.size else None,
    }


def build_tdms_groups(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    tdms_path = Path(manifest["tdms_path"])
    groups: list[dict[str, Any]] = []
    with TdmsFile.open(tdms_path) as tdms:
        for group in tdms.groups():
            explicit_group_time = detect_group_time_axis(tdms, group.name)
            plot_channels: list[dict[str, Any]] = []
            scalar_channels: list[dict[str, Any]] = []
            for channel in group.channels():
                values = np.asarray(channel[:])
                if values.dtype.kind not in "iufb":
                    continue
                numeric = np.asarray(values, dtype=float)
                unit = str(channel.properties.get("unit_string") or channel.properties.get("unit") or "").strip()
                if numeric.size <= 1:
                    scalar_channels.append(build_scalar_channel(channel.name, unit, numeric))
                    continue
                x_values, x_label = resolve_channel_x(channel, explicit_group_time)
                plot_channels.append(build_plot_channel(channel.name, unit, x_values, numeric, x_label))
            groups.append(
                {
                    "name": group.name,
                    "source_kind": "tdms_group",
                    "description": f"Source TDMS group with {len(group.channels())} channels.",
                    "plot_channels": plot_channels,
                    "scalar_channels": scalar_channels,
                    "channel_count": len(group.channels()),
                }
            )
    return groups


def build_official_group(case_root: Path) -> dict[str, Any]:
    official = pd.read_parquet(case_root / "official_known_families_wide.parquet")
    plot_channels: list[dict[str, Any]] = []
    time_values = official["time_s"].to_numpy(dtype=float)
    for column in official.columns:
        if column == "time_s":
            continue
        series = official[column].to_numpy(dtype=float)
        unit = "g" if column.endswith("_accel_g") else "mm"
        plot_channels.append(build_plot_channel(column, unit, time_values, series, "time_s"))
    return {
        "name": "Official Known Layer",
        "source_kind": "derived_preprocessing_group",
        "description": "Repository-standard layer limited to IIHS-explicit channel families.",
        "plot_channels": plot_channels,
        "scalar_channels": [],
        "channel_count": len(plot_channels),
    }


def build_t0_group(case_root: Path, manifest: dict[str, Any]) -> dict[str, Any]:
    t0 = pd.read_parquet(case_root / "exploratory_vehicle_longitudinal_t0_proxy.parquet")
    metrics = manifest["t0_proxy_assessment"]["metrics"]
    plot_channels = [
        build_plot_channel(
            "official_vehicle_longitudinal_basis",
            "g",
            t0["source_time_s"].to_numpy(dtype=float),
            t0["source_vehicle_longitudinal_accel_g"].to_numpy(dtype=float),
            "time_s",
        ),
        build_plot_channel(
            "t0_proxy_vehicle_longitudinal",
            "g",
            t0["shifted_time_s"].to_numpy(dtype=float),
            t0["vehicle_longitudinal_accel_g_t0_proxy"].to_numpy(dtype=float),
            "time_s",
        ),
    ]
    scalar_channels = [
        {
            "name": "t0_proxy_metrics",
            "unit": "",
            "values": [
                f"detected_bias_g={metrics['detected_bias_g']:.6f}",
                f"anchor_time_s={metrics['anchor_time_s']:.6f}",
                f"t0_time_s={metrics['t0_time_s']:.6f}",
                f"algorithm_mode={metrics['algorithm_mode']}",
            ],
            "sample_count": 4,
            "min": None,
            "max": None,
            "mean": None,
        }
    ]
    return {
        "name": "Exploratory T0 Proxy",
        "source_kind": "derived_exploratory_group",
        "description": "Comparison-only layer showing prior-project T0 alignment against the official longitudinal basis.",
        "plot_channels": plot_channels,
        "scalar_channels": scalar_channels,
        "channel_count": 2,
    }


def relative_plot_paths(case_name: str, dashboard_dir: Path, plots_root: Path) -> dict[str, str]:
    case_plot_root = plots_root / case_name
    overview = case_plot_root / "01_official_overview.png"
    detail = case_plot_root / "02_longitudinal_detail.png"
    assets_dir = dashboard_dir / "assets"
    assets_dir.mkdir(parents=True, exist_ok=True)

    def copy_asset(path: Path) -> str:
        if not path.exists():
            return ""
        target = assets_dir / path.name
        if not target.exists():
            shutil.copy2(path, target)
        return Path(__import__("os").path.relpath(target, start=dashboard_dir)).as_posix()

    return {
        "overview_rel": copy_asset(overview),
        "detail_rel": copy_asset(detail),
    }


def relative_pdf_catalog_path(dashboard_dir: Path, output_root: Path) -> str:
    catalog = output_root / "pdf_catalog" / "index.html"
    if not catalog.exists():
        return ""
    return Path(__import__("os").path.relpath(catalog, start=dashboard_dir)).as_posix()


def relative_excel_catalog_path(dashboard_dir: Path, output_root: Path) -> str:
    catalog = output_root / "excel_catalog" / "index.html"
    if not catalog.exists():
        return ""
    return Path(__import__("os").path.relpath(catalog, start=dashboard_dir)).as_posix()


def load_pdf_case_results(test_code: str, research_db_path: Path) -> dict[str, Any]:
    empty = {
        "available": False,
        "document_count": 0,
        "table_count": 0,
        "row_count": 0,
        "review_row_count": 0,
        "documents": [],
        "type_summary": [],
        "tables": [],
    }
    if not research_db_path.exists():
        return empty

    connection = sqlite3.connect(research_db_path)
    connection.row_factory = sqlite3.Row

    document_rows = connection.execute(
        """
        SELECT pdi.pdf_document_id,
               pdi.pdf_role,
               pdi.family_label,
               pdi.filename,
               pdi.local_path,
               pdi.page_count,
               COUNT(DISTINCT prt.pdf_result_table_id) AS table_count,
               COUNT(prr.pdf_result_row_id) AS row_count
          FROM pdf_document_inventory pdi
          LEFT JOIN pdf_result_tables prt ON prt.pdf_document_id = pdi.pdf_document_id
          LEFT JOIN pdf_result_rows prr ON prr.pdf_result_table_id = prt.pdf_result_table_id
         WHERE pdi.test_code = ?
         GROUP BY pdi.pdf_document_id, pdi.pdf_role, pdi.family_label, pdi.filename, pdi.local_path, pdi.page_count
         ORDER BY pdi.pdf_role, pdi.filename
        """,
        (test_code,),
    ).fetchall()

    type_rows = connection.execute(
        """
        SELECT table_type,
               COUNT(DISTINCT pdf_result_table_id) AS table_count,
               COUNT(*) AS row_count,
               SUM(CASE WHEN quality_status = 'review' THEN 1 ELSE 0 END) AS review_row_count
          FROM pdf_result_row_catalog
         WHERE test_code = ?
         GROUP BY table_type
         ORDER BY row_count DESC, table_type
        """,
        (test_code,),
    ).fetchall()

    table_rows = connection.execute(
        """
        SELECT pdf_result_table_id,
               pdf_document_id,
               pdf_role,
               family_label,
               document_title,
               local_path,
               page_number,
               table_order,
               table_ref,
               table_title,
               table_type,
               table_group,
               COUNT(*) AS row_count,
               SUM(CASE WHEN quality_status = 'review' THEN 1 ELSE 0 END) AS review_row_count
          FROM pdf_result_row_catalog
         WHERE test_code = ?
         GROUP BY pdf_result_table_id, pdf_document_id, pdf_role, family_label, document_title, local_path, page_number, table_order, table_ref, table_title, table_type, table_group
         ORDER BY page_number, table_order
        """,
        (test_code,),
    ).fetchall()

    row_rows = connection.execute(
        """
        SELECT pdf_result_table_id,
               row_order,
               section_name,
               seat_position,
               label,
               code,
               normalized_label,
               quality_status,
               quality_score,
               threshold_text,
               result_text,
               time_text,
               left_text,
               left_time_text,
               right_text,
               right_time_text,
               longitudinal_text,
               lateral_text,
               vertical_text,
               resultant_text,
               measure_text,
               unit
          FROM pdf_result_row_catalog
         WHERE test_code = ?
         ORDER BY page_number, table_order, row_order
        """,
        (test_code,),
    ).fetchall()
    connection.close()

    rows_by_table: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in row_rows:
        payload = {
            "row_order": row["row_order"],
            "section_name": row["section_name"] or "",
            "seat_position": row["seat_position"] or "",
            "label": row["label"] or "",
            "code": row["code"] or "",
            "normalized_label": row["normalized_label"] or "",
            "quality_status": row["quality_status"] or "",
            "quality_score": round(float(row["quality_score"]), 3) if row["quality_score"] is not None else None,
            "threshold_text": row["threshold_text"] or "",
            "result_text": row["result_text"] or "",
            "time_text": row["time_text"] or "",
            "left_text": row["left_text"] or "",
            "left_time_text": row["left_time_text"] or "",
            "right_text": row["right_text"] or "",
            "right_time_text": row["right_time_text"] or "",
            "longitudinal_text": row["longitudinal_text"] or "",
            "lateral_text": row["lateral_text"] or "",
            "vertical_text": row["vertical_text"] or "",
            "resultant_text": row["resultant_text"] or "",
            "measure_text": row["measure_text"] or "",
            "unit": row["unit"] or "",
        }
        rows_by_table[row["pdf_result_table_id"]].append(payload)

    documents = [dict(row) for row in document_rows]
    documents_by_id = {row["pdf_document_id"]: row for row in documents}
    type_summary = [dict(row) for row in type_rows]
    tables = []
    for row in table_rows:
        payload = dict(row)
        document = documents_by_id.get(row["pdf_document_id"], {})
        local_path = row["local_path"] or ""
        payload["filename"] = document.get("filename") or (Path(local_path).name if local_path else "") or row["document_title"] or ""
        payload["rows"] = rows_by_table.get(row["pdf_result_table_id"], [])
        tables.append(payload)

    return {
        "available": bool(tables),
        "document_count": len(documents),
        "table_count": len(tables),
        "row_count": sum(int(row["row_count"]) for row in tables),
        "review_row_count": sum(int(row["review_row_count"] or 0) for row in tables),
        "documents": documents,
        "type_summary": type_summary,
        "tables": tables,
    }


def clean_excel_chart_label(value: Any) -> str:
    text = str(value or "").replace("\r", "\n").strip()
    lines = [segment.strip() for segment in text.splitlines() if segment and segment.strip()]
    return "\n".join(lines)


def parse_chart_number(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text or text in {"--", "-", "n/a", "N/A"}:
        return None
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def normalize_chart_label(value: Any) -> str:
    text = str(value or "").replace("\r", " ").replace("\n", " ").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def load_excel_intrusion_chart(test_code: str, research_db_path: Path) -> dict[str, Any]:
    empty = {
        "available": False,
        "source_filename": "",
        "source_path": "",
        "source_sheet": "",
        "title": "",
        "axis_max_cm": 0,
        "rows": [],
    }
    if not research_db_path.exists():
        return empty

    connection = sqlite3.connect(research_db_path)
    connection.row_factory = sqlite3.Row
    workbook_row = connection.execute(
        """
        SELECT filename, local_path
          FROM excel_workbook_inventory
         WHERE test_code = ? AND workbook_type = 'intrusion' AND local_path IS NOT NULL
         ORDER BY filename
         LIMIT 1
        """,
        (test_code,),
    ).fetchone()
    connection.close()
    if workbook_row is None:
        return empty

    workbook_path = Path(workbook_row["local_path"])
    if not workbook_path.exists():
        return empty

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet_name = test_code if test_code in workbook.sheetnames else next(
            (name for name in workbook.sheetnames if str(name).upper().startswith(test_code.upper())),
            "",
        )
        if not sheet_name:
            return empty

        worksheet = workbook[sheet_name]
        title = str(worksheet["F1"].value or f"{test_code} Intrusion").strip()
        header_row = None
        max_scan_row = min(int(worksheet.max_row or 0), 120)
        for row_idx in range(1, max_scan_row + 1):
            header_value = worksheet.cell(row=row_idx, column=15).value
            if isinstance(header_value, str) and "resultant" in header_value.lower():
                header_row = row_idx
                break
        if header_row is None:
            return empty

        rows: list[dict[str, Any]] = []
        blank_count = 0
        for row_idx in range(header_row + 1, min(header_row + 24, int(worksheet.max_row or header_row)) + 1):
            raw_label = worksheet.cell(row=row_idx, column=14).value
            if raw_label in (None, ""):
                blank_count += 1
                if rows and blank_count >= 2:
                    break
                continue
            blank_count = 0

            measured = worksheet.cell(row=row_idx, column=15).value
            good = worksheet.cell(row=row_idx, column=16).value
            acceptable = worksheet.cell(row=row_idx, column=17).value
            marginal = worksheet.cell(row=row_idx, column=18).value
            if all(value in (None, "") for value in (measured, good, acceptable, marginal)):
                continue

            label = clean_excel_chart_label(raw_label)
            measured_cm = float(measured) if measured not in (None, "") else None
            good_cm = float(good) if good not in (None, "") else 0.0
            acceptable_band_cm = float(acceptable) if acceptable not in (None, "") else 0.0
            marginal_band_cm = float(marginal) if marginal not in (None, "") else 0.0
            acceptable_upper_cm = good_cm + acceptable_band_cm
            marginal_upper_cm = acceptable_upper_cm + marginal_band_cm
            if measured_cm is None:
                rating = "Not measured"
            elif measured_cm <= good_cm:
                rating = "Good"
            elif measured_cm <= acceptable_upper_cm:
                rating = "Acceptable"
            elif measured_cm <= marginal_upper_cm:
                rating = "Marginal"
            else:
                rating = "Poor"

            rows.append(
                {
                    "label": label.replace("\n", " ").strip(),
                    "label_html": label.replace("\n", "<br>"),
                    "measured_cm": round(measured_cm, 3) if measured_cm is not None else None,
                    "good_cm": round(good_cm, 3),
                    "acceptable_upper_cm": round(acceptable_upper_cm, 3),
                    "marginal_upper_cm": round(marginal_upper_cm, 3),
                    "rating": rating,
                }
            )
        if not rows:
            return empty

        max_threshold = max(row["marginal_upper_cm"] for row in rows)
        max_measured = max((row["measured_cm"] or 0.0) for row in rows)
        axis_seed = max(max_threshold * 1.5, max_measured * 1.25, 15.0)
        axis_max_cm = int(math.ceil(axis_seed / 5.0) * 5)
        return {
            "available": True,
            "source_filename": workbook_row["filename"] or workbook_path.name,
            "source_path": str(workbook_path),
            "source_sheet": sheet_name,
            "title": title,
            "axis_max_cm": axis_max_cm,
            "rows": rows,
        }
    finally:
        workbook.close()


def build_pdf_intrusion_overlay(pdf_results: dict[str, Any], excel_intrusion_chart: dict[str, Any]) -> dict[str, Any]:
    empty = {
        "available": False,
        "matched_count": 0,
        "row_count": 0,
        "source_table_title": "",
        "rows": [],
    }
    if not excel_intrusion_chart.get("available"):
        return empty

    intrusion_table = next((row for row in pdf_results.get("tables", []) if row.get("table_type") == "intrusion"), None)
    if intrusion_table is None:
        return empty

    normalized_rows = []
    for row in intrusion_table.get("rows", []):
        normalized_rows.append(
            {
                "norm": normalize_chart_label(row.get("label")),
                "row": row,
            }
        )

    def filter_rows(prefix: str) -> list[dict[str, Any]]:
        return [entry["row"] for entry in normalized_rows if entry["norm"].startswith(prefix)]

    def first_row(*terms: str) -> dict[str, Any] | None:
        for entry in normalized_rows:
            if all(term in entry["norm"] for term in terms):
                return entry["row"]
        return None

    def value_from_row(row: dict[str, Any] | None, field: str, *, absolute: bool = False) -> float | None:
        if row is None:
            return None
        numeric = parse_chart_number(row.get(field))
        if numeric is None:
            return None
        return abs(numeric) if absolute else numeric

    def max_resultant(prefix: str) -> float | None:
        values = [parse_chart_number(row.get("resultant_text")) for row in filter_rows(prefix)]
        values = [value for value in values if value is not None]
        return max(values) if values else None

    extracted_values = {
        "Lower Hinge Pillar": max_resultant("lower hinge pillar z"),
        "Footrest": value_from_row(first_row("footrest"), "resultant_text"),
        "Left Toepan": value_from_row(first_row("left", "toepan"), "resultant_text"),
        "Brake Pedal": value_from_row(first_row("brake pedal"), "resultant_text"),
        "Parking Brake Pedal": value_from_row(first_row("parking brake"), "resultant_text"),
        "Rocker Panel (lat)": value_from_row(first_row("rocker panel", "average lateral"), "lateral_text", absolute=True),
        "Steering Column (long)": value_from_row(first_row("steering column"), "longitudinal_text", absolute=True),
        "Upper Hinge Pillar": max_resultant("upper hinge pillar z"),
        "Upper Dash": value_from_row(first_row("upper dash"), "resultant_text"),
        "Left Instrument Panel": value_from_row(first_row("left", "instrument panel"), "resultant_text"),
    }

    overlay_rows: list[dict[str, Any]] = []
    matched_count = 0
    for base_row in excel_intrusion_chart.get("rows", []):
        pdf_value = extracted_values.get(base_row["label"])
        if pdf_value is not None:
            matched_count += 1
            if pdf_value <= base_row["good_cm"]:
                rating = "Good"
            elif pdf_value <= base_row["acceptable_upper_cm"]:
                rating = "Acceptable"
            elif pdf_value <= base_row["marginal_upper_cm"]:
                rating = "Marginal"
            else:
                rating = "Poor"
        else:
            rating = "Not measured"
        overlay_rows.append(
            {
                "label": base_row["label"],
                "pdf_measured_cm": round(pdf_value, 3) if pdf_value is not None else None,
                "pdf_rating": rating,
            }
        )

    return {
        "available": matched_count > 0,
        "matched_count": matched_count,
        "row_count": len(overlay_rows),
        "source_table_title": intrusion_table.get("table_title", ""),
        "rows": overlay_rows,
    }


def build_dashboard_data(case_root: Path, dashboard_dir: Path, plots_root: Path, output_root: Path, research_db_path: Path) -> dict[str, Any]:
    manifest = load_manifest(case_root)
    groups = [build_official_group(case_root), build_t0_group(case_root, manifest), *build_tdms_groups(manifest)]
    excel_intrusion_chart = load_excel_intrusion_chart(manifest["test_code"], research_db_path)
    pdf_results = load_pdf_case_results(manifest["test_code"], research_db_path)
    excel_intrusion_chart["pdf_overlay"] = build_pdf_intrusion_overlay(pdf_results, excel_intrusion_chart)
    return {
        "case_name": case_root.name,
        "filegroup_id": manifest["filegroup_id"],
        "test_code": manifest["test_code"],
        "vehicle_make_model": manifest["vehicle_make_model"],
        "tdms_path": manifest["tdms_path"],
        "time_basis": manifest["time_basis"],
        "official_policy": manifest["official_policy"],
        "t0_proxy_assessment": manifest["t0_proxy_assessment"],
        "plots": relative_plot_paths(case_root.name, dashboard_dir, plots_root),
        "pdf_catalog_rel": relative_pdf_catalog_path(dashboard_dir, output_root),
        "excel_catalog_rel": relative_excel_catalog_path(dashboard_dir, output_root),
        "excel_intrusion_chart": excel_intrusion_chart,
        "pdf_results": pdf_results,
        "groups": groups,
        "summary": {
            "dashboard_group_count": len(groups),
            "tdms_group_count": sum(1 for group in groups if group["source_kind"] == "tdms_group"),
            "derived_group_count": sum(1 for group in groups if group["source_kind"] != "tdms_group"),
            "official_channel_count": groups[0]["channel_count"],
        },
    }


def dashboard_html(data: dict[str, Any]) -> str:
    data_json = json.dumps(data, ensure_ascii=False).replace("</script>", "<\\/script>")
    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{data["test_code"]} Signal Dashboard</title>
  <link rel="icon" href="data:,">
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;700&family=IBM+Plex+Sans:wght@400;500;600&display=swap" rel="stylesheet">
  <script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
  <style>
    :root {{
      --bg:#f3ede3; --card:#fffaf2; --muted:#6a6054; --ink:#1d1814; --line:rgba(0,0,0,.09);
      --accent:#bc4b32; --accent2:#0f7173; --accent3:#785589;
    }}
    * {{ box-sizing:border-box; }}
    body {{
      margin:0; color:var(--ink); font-family:"IBM Plex Sans","Segoe UI",sans-serif;
      background:
        radial-gradient(circle at top left, rgba(188,75,50,.10), transparent 28%),
        radial-gradient(circle at top right, rgba(15,113,115,.10), transparent 32%),
        linear-gradient(180deg, #f8f3ea 0%, #f3ede3 100%);
    }}
    .page {{ max-width:1560px; margin:0 auto; padding:26px; }}
    .card {{
      background:rgba(255,250,242,.94); border:1px solid var(--line); border-radius:22px;
      box-shadow:0 12px 32px rgba(29,24,20,.08); padding:20px 22px;
    }}
    .hero {{ display:grid; grid-template-columns:1.15fr .85fr; gap:18px; }}
    .eyebrow {{
      font-family:"Space Grotesk",sans-serif; text-transform:uppercase; letter-spacing:.14em;
      font-size:12px; color:var(--muted); margin-bottom:10px;
    }}
    h1,h2,h3 {{ font-family:"Space Grotesk",sans-serif; margin:0; line-height:1.06; }}
    h1 {{ font-size:clamp(32px,4vw,54px); margin-bottom:12px; }}
    .hero-copy {{ color:#322a22; line-height:1.55; font-size:16px; }}
    .meta-grid {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; margin-top:18px; }}
    .meta-pill {{ background:#f5edde; border:1px solid var(--line); border-radius:16px; padding:12px 14px; }}
    .meta-pill span {{ display:block; color:var(--muted); font-size:12px; text-transform:uppercase; letter-spacing:.08em; margin-bottom:5px; }}
    .meta-pill strong {{ font-size:18px; }}
    .note {{ background:linear-gradient(135deg, rgba(188,75,50,.08), rgba(15,113,115,.05)); border-radius:18px; border:1px solid var(--line); padding:16px; }}
    .note + .note {{ margin-top:12px; }}
    .note strong {{ display:block; margin-bottom:8px; font-family:"Space Grotesk",sans-serif; font-size:18px; }}
    .note a {{ color:var(--accent2); text-decoration:none; font-weight:600; }}
    .gallery {{ margin-top:18px; display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:18px; }}
    .gallery img {{ display:block; width:100%; border-radius:18px; border:1px solid var(--line); background:#fff; }}
    .gallery figcaption {{ margin-top:10px; font-size:14px; color:var(--muted); }}
    .browser {{ margin-top:20px; display:grid; grid-template-columns:320px minmax(0,1fr); gap:18px; }}
    .sidebar {{ position:sticky; top:18px; }}
    .nav-row,.action-row {{ display:flex; gap:8px; }}
    .nav-row button,.action-row button,.group-btn {{
      border:0; border-radius:14px; cursor:pointer; transition:transform .14s ease, background .14s ease;
    }}
    .nav-row button,.action-row button {{ flex:1; padding:10px 12px; background:#241d17; color:#f8f3ea; font-weight:600; }}
    .nav-row button:hover,.action-row button:hover,.group-btn:hover {{ transform:translateY(-1px); }}
    .group-list {{ display:grid; gap:8px; max-height:60vh; overflow:auto; margin-top:12px; }}
    .group-btn {{ text-align:left; padding:13px 14px; background:#f5edde; border:1px solid transparent; }}
    .group-btn.active {{ background:#241d17; color:#fff8ef; }}
    .group-btn small {{ display:block; margin-top:4px; opacity:.72; }}
    .main {{ display:grid; gap:18px; }}
    .group-header {{ display:grid; grid-template-columns:minmax(0,1fr) auto; gap:14px; align-items:start; }}
    .group-subtitle,.muted {{ color:var(--muted); line-height:1.5; }}
    .chips {{ display:flex; gap:10px; flex-wrap:wrap; justify-content:flex-end; }}
    .chip {{ background:#f5edde; border:1px solid var(--line); border-radius:14px; padding:9px 11px; font-size:13px; color:var(--muted); }}
    .controls {{ display:grid; grid-template-columns:320px minmax(0,1fr); gap:18px; }}
    .search input {{
      width:100%; padding:12px 14px; border-radius:14px; border:1px solid var(--line); background:#fffdf8; font:inherit;
    }}
    .channel-list {{ display:grid; gap:8px; max-height:560px; overflow:auto; margin-top:12px; }}
    .channel-row {{
      display:grid; grid-template-columns:auto 1fr auto; gap:10px; align-items:center;
      padding:10px 12px; border-radius:14px; background:#fffdf8; border:1px solid var(--line);
    }}
    .channel-row input {{ width:16px; height:16px; }}
    .channel-meta {{ font-size:12px; color:var(--muted); }}
    .stats-table {{ width:100%; border-collapse:collapse; font-size:14px; }}
    .stats-table th,.stats-table td {{ text-align:left; padding:10px 12px; border-bottom:1px solid var(--line); vertical-align:top; }}
    .stats-table th {{ font-size:12px; text-transform:uppercase; letter-spacing:.08em; color:var(--muted); }}
    .table-shell {{ max-height:520px; overflow:auto; border:1px solid var(--line); border-radius:18px; background:#fffdf8; }}
    .excel-results,.pdf-results {{ margin-top:20px; display:grid; grid-template-columns:1fr; gap:18px; }}
    .excel-chart-grid {{ display:grid; grid-template-columns:minmax(0,1.2fr) minmax(340px,.8fr); gap:18px; align-items:start; margin-top:14px; }}
    .pdf-summary-grid {{ display:grid; grid-template-columns:minmax(0,1.15fr) minmax(340px,.85fr); gap:18px; align-items:start; margin-top:14px; }}
    .pdf-table-browser {{ display:grid; grid-template-columns:300px minmax(0,1fr); gap:18px; align-items:start; }}
    .pdf-table-list {{ display:grid; gap:8px; max-height:560px; overflow:auto; }}
    .pdf-table-btn {{ text-align:left; padding:13px 14px; border:1px solid transparent; border-radius:14px; background:#f5edde; cursor:pointer; transition:transform .14s ease, background .14s ease; }}
    .pdf-table-btn:hover {{ transform:translateY(-1px); }}
    .pdf-table-btn.active {{ background:#241d17; color:#fff8ef; }}
    .pdf-table-btn.active .muted {{ color:rgba(255,248,239,.72); }}
    .pdf-table-head {{ display:flex; justify-content:space-between; align-items:flex-start; gap:12px; flex-wrap:wrap; margin-bottom:12px; }}
    .pdf-empty {{ display:flex; align-items:center; justify-content:center; min-height:220px; color:var(--muted); font-size:15px; }}
    .rating-pill {{ display:inline-flex; align-items:center; border-radius:999px; padding:5px 10px; font-size:12px; font-weight:700; letter-spacing:.04em; text-transform:uppercase; }}
    .rating-good {{ background:#17f200; color:#11330b; }}
    .rating-acceptable {{ background:#f2ef00; color:#4a4700; }}
    .rating-marginal {{ background:#ff7a00; color:#fff4e8; }}
    .rating-poor {{ background:#f51610; color:#fff5f4; }}
    .rating-unrated {{ background:#e8ded0; color:#6a6054; }}
    .scalar-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:12px; margin-top:14px; }}
    .scalar-card {{ padding:14px 16px; border-radius:16px; background:#f5edde; border:1px solid var(--line); }}
    .scalar-card strong {{ display:block; margin-bottom:8px; font-family:"Space Grotesk",sans-serif; }}
    .scalar-card code {{ display:block; white-space:pre-wrap; font-family:ui-monospace,Consolas,monospace; font-size:12px; line-height:1.5; }}
    .footnote {{ margin-top:8px; color:var(--muted); font-size:13px; }}
    @media (max-width:1180px) {{
      .hero,.gallery,.browser,.controls,.excel-results,.excel-chart-grid,.pdf-results,.pdf-summary-grid,.pdf-table-browser {{ grid-template-columns:1fr; }}
      .sidebar {{ position:static; }}
      .meta-grid {{ grid-template-columns:repeat(2,minmax(0,1fr)); }}
    }}
  </style>
</head>
<body>
  <div class="page">
    <section class="hero">
      <div class="card">
        <div class="eyebrow">IIHS Small Overlap Dashboard</div>
        <h1>{data["test_code"]} · {data["vehicle_make_model"]}</h1>
        <div class="hero-copy">
          전처리 산출물과 TDMS 전체 그룹을 브라우저에서 바로 넘겨볼 수 있도록 만든 정적 대시보드입니다.
          상단은 공식 전처리 개요, 하단은 <strong>Official Known Layer</strong>, <strong>Exploratory T0 Proxy</strong>,
          그리고 TDMS의 모든 그룹을 인터랙티브하게 탐색합니다.
        </div>
        <div class="meta-grid">
          <div class="meta-pill"><span>Filegroup</span><strong>{data["filegroup_id"]}</strong></div>
          <div class="meta-pill"><span>Dashboard Groups</span><strong>{data["summary"]["dashboard_group_count"]}</strong></div>
          <div class="meta-pill"><span>Official Channels</span><strong>{data["summary"]["official_channel_count"]}</strong></div>
          <div class="meta-pill"><span>Sample Rate</span><strong>{data["time_basis"]["sample_rate_hz"]} Hz</strong></div>
        </div>
      </div>
      <div class="card">
        <div class="note">
          <strong>Time Basis</strong>
          <div>{data["time_basis"]["policy"]}</div>
          <div class="footnote">Selected source: {data["time_basis"]["selected_source"]}</div>
          <div class="footnote">Axis note: {data["time_basis"]["note"] or "n/a"}</div>
        </div>
        <div class="note">
          <strong>T0 Proxy</strong>
          <div>Detected bias: {data["t0_proxy_assessment"]["metrics"]["detected_bias_g"]:.3f} g</div>
          <div>Anchor: {data["t0_proxy_assessment"]["metrics"]["anchor_time_s"]*1000:.2f} ms</div>
          <div>T0 proxy: {data["t0_proxy_assessment"]["metrics"]["t0_time_s"]*1000:.2f} ms</div>
          <div class="footnote">Status: {data["t0_proxy_assessment"]["status"]}</div>
        </div>
        {("<div class='note'><strong>Research DB</strong><div>전체 PDF 전수조사 결과를 카탈로그 대시보드로 연결합니다.</div>" + ("<div class='footnote'><a href='" + data["pdf_catalog_rel"] + "'>Open PDF catalog dashboard</a></div>" if data["pdf_catalog_rel"] else "") + ("<div class='footnote'><a href='" + data["excel_catalog_rel"] + "'>Open Excel catalog dashboard</a></div>" if data["excel_catalog_rel"] else "") + "</div>") if data["pdf_catalog_rel"] or data["excel_catalog_rel"] else ""}
      </div>
    </section>

    <section class="gallery">
      <figure class="card">
        <h2 style="font-size:24px; margin-bottom:14px;">Official Overview</h2>
        <img src="{data["plots"]["overview_rel"]}" alt="Official preprocessing overview" />
        <figcaption>공식 레이어의 vehicle, seat, foot 채널 묶음.</figcaption>
      </figure>
      <figure class="card">
        <h2 style="font-size:24px; margin-bottom:14px;">Longitudinal Detail</h2>
        <img src="{data["plots"]["detail_rel"]}" alt="Longitudinal detail comparison" />
        <figcaption>Longitudinal raw reference, official basis, exploratory T0 proxy 비교.</figcaption>
      </figure>
    </section>

    <section class="excel-results">
      <div class="card">
        <div class="pdf-table-head">
          <div>
            <div class="eyebrow">Excel Result Chart</div>
            <h2 style="font-size:28px;">Intrusion Rating Envelope</h2>
            <div class="muted" id="excel-intrusion-subtitle"></div>
          </div>
          <div class="chips" id="excel-intrusion-chips"></div>
        </div>
        <div class="excel-chart-grid">
          <div>
            <div id="excel-intrusion-plot" style="width:100%; height:520px;"></div>
          </div>
          <div>
            <div class="footnote" id="excel-intrusion-note"></div>
            <div class="table-shell" style="max-height:520px; margin-top:10px;">
              <table class="stats-table">
                <thead>
                  <tr><th>Location</th><th>Excel (cm)</th><th>PDF (cm)</th><th>Delta</th><th>Good</th><th>Accept.</th><th>Marginal</th><th>PDF Rating</th></tr>
                </thead>
                <tbody id="excel-intrusion-body"></tbody>
              </table>
            </div>
          </div>
        </div>
      </div>
    </section>

    <section class="pdf-results">
      <div class="card">
        <div class="pdf-table-head">
          <div>
            <div class="eyebrow">PDF Extracted Results</div>
            <h2 style="font-size:28px;">Parsed Result Type Summary</h2>
            <div class="muted">This case report PDF is parsed into injury, intrusion, dummy clearance, and restraint timing tables so the extracted rows can be reviewed without leaving the signal dashboard.</div>
          </div>
          <div class="chips" id="pdf-summary-kpis" style="justify-content:flex-start;"></div>
        </div>
        <div class="pdf-summary-grid">
          <div>
            <div id="pdf-type-plot" style="width:100%; height:420px;"></div>
          </div>
          <div>
            <div class="footnote" id="pdf-summary-note"></div>
            <div class="table-shell" style="max-height:420px; margin-top:10px;">
              <table class="stats-table">
                <thead>
                  <tr><th>Type</th><th>Tables</th><th>Rows</th><th>Review rows</th></tr>
                </thead>
                <tbody id="pdf-type-body"></tbody>
              </table>
            </div>
          </div>
        </div>
      </div>
      <div class="card">
        <div class="pdf-table-head">
          <div>
            <div class="eyebrow">PDF Table Browser</div>
            <h2 style="font-size:28px;" id="pdf-table-title">Parsed Table Detail</h2>
            <div class="muted" id="pdf-table-subtitle"></div>
          </div>
          <div class="chips" id="pdf-table-chips"></div>
        </div>
        <div class="pdf-table-browser">
          <div class="pdf-table-list" id="pdf-table-list"></div>
          <div>
            <div class="table-shell">
              <table class="stats-table">
                <thead><tr id="pdf-table-head-row"></tr></thead>
                <tbody id="pdf-table-body"></tbody>
              </table>
            </div>
            <div class="footnote" id="pdf-table-note"></div>
          </div>
        </div>
      </div>
    </section>

    <section class="browser">
      <aside class="card sidebar">
        <h2 style="font-size:24px;">Group Browser</h2>
        <div class="nav-row" style="margin-top:12px;">
          <button id="prev-group" type="button">Previous</button>
          <button id="next-group" type="button">Next</button>
        </div>
        <div class="footnote" id="group-counter"></div>
        <div class="group-list" id="group-list"></div>
      </aside>
      <main class="main">
        <section class="card">
          <div class="group-header">
            <div>
              <div class="eyebrow" id="group-kind"></div>
              <h2 id="group-title" style="font-size:32px;"></h2>
              <div class="group-subtitle" id="group-description"></div>
            </div>
            <div class="chips" id="group-chips"></div>
          </div>
        </section>
        <section class="card">
          <div class="controls">
            <div>
              <div class="search"><input id="channel-filter" type="text" placeholder="채널 이름 검색" /></div>
              <div class="action-row" style="margin-top:10px;">
                <button id="select-default" type="button">Default</button>
                <button id="select-visible" type="button">Select Visible</button>
                <button id="clear-selection" type="button">Clear</button>
              </div>
              <div class="footnote" id="selection-note" style="margin-top:12px;"></div>
              <div class="channel-list" id="channel-list"></div>
            </div>
            <div><div id="group-plot" style="width:100%; height:620px;"></div></div>
          </div>
        </section>
        <section class="card">
          <h3 style="font-size:22px; margin-bottom:12px;">Selected Channel Stats</h3>
          <table class="stats-table">
            <thead>
              <tr>
                <th>Channel</th><th>Unit</th><th>Points</th><th>Mean</th><th>Min</th><th>Max</th><th>Axis</th>
              </tr>
            </thead>
            <tbody id="stats-body"></tbody>
          </table>
          <div class="scalar-grid" id="scalar-grid"></div>
          <div class="footnote">브라우저 성능을 위해 plot은 최대 {MAX_PLOT_POINTS} 포인트로 downsample 했고, 통계는 원본 sample 기준입니다.</div>
        </section>
      </main>
    </section>
  </div>
  <script id="dashboard-data" type="application/json">{data_json}</script>
  <script>
    const dashboard = JSON.parse(document.getElementById("dashboard-data").textContent);
    const palette = ["#bc4b32","#0f7173","#785589","#c97c10","#2f6690","#6d9f71","#ba6f8b","#495057","#7f5539","#386641"];
    const pdfTypeLabels = {{
      head_injury: "Head Injury",
      neck_injury: "Neck Injury",
      chest_injury: "Chest Injury",
      thigh_hip_injury: "Thigh and Hip Injury",
      leg_foot_injury: "Leg and Foot Injury",
      intrusion: "Intrusion",
      dummy_clearance: "Dummy Clearance",
      restraint_kinematics: "Restraint and Kinematics",
    }};
    const state = {{ groupIndex: 0, filter: "", selections: {{}}, pdfTableId: dashboard.pdf_results?.tables?.[0]?.pdf_result_table_id ?? null }};

    function currentGroup() {{ return dashboard.groups[state.groupIndex]; }}
    function esc(value) {{
      return String(value ?? "")
        .replace(/&/g, "&amp;")
        .replace(/</g, "&lt;")
        .replace(/>/g, "&gt;")
        .replace(/"/g, "&quot;");
    }}
    function displayValue(value) {{
      return value === undefined || value === null || value === "" ? "n/a" : String(value);
    }}
    function prettyPdfType(value) {{
      if (!value) return "Unknown";
      return pdfTypeLabels[value] || value.replace(/_/g, " ").replace(/\b\w/g, (char) => char.toUpperCase());
    }}
    function pdfResults() {{
      return dashboard.pdf_results || {{ available: false, document_count: 0, table_count: 0, row_count: 0, review_row_count: 0, type_summary: [], tables: [] }};
    }}
    function excelIntrusion() {{
      return dashboard.excel_intrusion_chart || {{ available: false, rows: [], axis_max_cm: 0 }};
    }}
    function selectedPdfTable() {{
      const tables = pdfResults().tables || [];
      if (!tables.length) return null;
      if (!tables.some((table) => table.pdf_result_table_id === state.pdfTableId)) state.pdfTableId = tables[0].pdf_result_table_id;
      return tables.find((table) => table.pdf_result_table_id === state.pdfTableId) || tables[0];
    }}
    function ratingClass(value) {{
      if (value === "Good") return "rating-pill rating-good";
      if (value === "Acceptable") return "rating-pill rating-acceptable";
      if (value === "Marginal") return "rating-pill rating-marginal";
      if (value === "Poor") return "rating-pill rating-poor";
      return "rating-pill rating-unrated";
    }}
    function defaultSelection(group) {{
      const names = group.plot_channels.map((channel) => channel.name);
      return group.name === "Official Known Layer" ? names : names.slice(0, Math.min(8, names.length));
    }}
    function ensureSelection(group) {{
      if (!state.selections[group.name]) state.selections[group.name] = new Set(defaultSelection(group));
      return state.selections[group.name];
    }}
    function filteredChannels(group) {{
      const q = state.filter.trim().toLowerCase();
      return q ? group.plot_channels.filter((channel) => channel.name.toLowerCase().includes(q)) : group.plot_channels;
    }}
    function renderGroupButtons() {{
      const list = document.getElementById("group-list");
      list.innerHTML = "";
      dashboard.groups.forEach((group, index) => {{
        const button = document.createElement("button");
        button.className = "group-btn" + (index === state.groupIndex ? " active" : "");
        button.type = "button";
        button.innerHTML = `<strong>${{group.name}}</strong><small>${{group.source_kind}} · ${{group.channel_count}} channels</small>`;
        button.onclick = () => {{
          state.groupIndex = index;
          state.filter = "";
          document.getElementById("channel-filter").value = "";
          render();
        }};
        list.appendChild(button);
      }});
      document.getElementById("group-counter").textContent = `Group ${{state.groupIndex + 1}} / ${{dashboard.groups.length}}`;
    }}
    function renderHeader(group) {{
      document.getElementById("group-kind").textContent = group.source_kind;
      document.getElementById("group-title").textContent = group.name;
      document.getElementById("group-description").textContent = group.description;
      const chips = document.getElementById("group-chips");
      chips.innerHTML = "";
      [
        `Plot channels: ${{group.plot_channels.length}}`,
        `Scalar channels: ${{group.scalar_channels.length}}`,
        `Total channels: ${{group.channel_count}}`,
      ].forEach((text) => {{
        const chip = document.createElement("div");
        chip.className = "chip";
        chip.textContent = text;
        chips.appendChild(chip);
      }});
    }}
    function renderChannelList(group) {{
      const selection = ensureSelection(group);
      const container = document.getElementById("channel-list");
      container.innerHTML = "";
      const visible = filteredChannels(group);
      document.getElementById("selection-note").textContent = `Selected ${{selection.size}} / ${{group.plot_channels.length}} channels`;
      visible.forEach((channel) => {{
        const row = document.createElement("label");
        row.className = "channel-row";
        row.innerHTML = `
          <input type="checkbox" ${{selection.has(channel.name) ? "checked" : ""}} />
          <div>
            <div><strong>${{channel.name}}</strong></div>
            <div class="channel-meta">${{channel.unit || "unit n/a"}} · ${{channel.stats.sample_count}} pts · ${{channel.stats.x_label}}</div>
          </div>
          <div class="channel-meta">min ${{channel.stats.min ?? "n/a"}}<br/>max ${{channel.stats.max ?? "n/a"}}</div>
        `;
        row.querySelector("input").onchange = (event) => {{
          if (event.target.checked) selection.add(channel.name);
          else selection.delete(channel.name);
          renderPlot(group);
          renderStats(group);
        }};
        container.appendChild(row);
      }});
      if (!visible.length) {{
        const empty = document.createElement("div");
        empty.className = "footnote";
        empty.textContent = "검색 조건에 맞는 채널이 없습니다.";
        container.appendChild(empty);
      }}
    }}
    function renderPlot(group) {{
      const selection = ensureSelection(group);
      const chosen = group.plot_channels.filter((channel) => selection.has(channel.name));
      const traces = chosen.map((channel, index) => ({{
        x: channel.x,
        y: channel.y,
        type: "scattergl",
        mode: "lines",
        name: channel.name,
        line: {{ width: 1.8, color: palette[index % palette.length] }},
        hovertemplate: "<b>%{{fullData.name}}</b><br>x=%{{x}}<br>y=%{{y}}<extra></extra>",
      }}));
      const layout = {{
        margin: {{ l: 60, r: 18, t: 28, b: 54 }},
        paper_bgcolor: "#fffaf2",
        plot_bgcolor: "#fffdf8",
        font: {{ family: "IBM Plex Sans, sans-serif", color: "#1d1814" }},
        xaxis: {{ title: chosen.some((row) => row.x_label === "time_s") ? "Time (s)" : "Sample Index", gridcolor: "rgba(0,0,0,.09)" }},
        yaxis: {{ title: chosen.length === 1 ? (chosen[0].unit || "value") : "Value", gridcolor: "rgba(0,0,0,.09)" }},
        legend: {{ orientation: "h", y: 1.15 }},
        hovermode: "closest",
        annotations: traces.length ? [] : [{{
          text: "선택된 채널이 없습니다.",
          x: .5, y: .5, xref: "paper", yref: "paper", showarrow: false, font: {{ size: 18, color: "#6a6054" }}
        }}],
      }};
      Plotly.react("group-plot", traces, layout, {{ responsive: true, displaylogo: false, modeBarButtonsToRemove: ["lasso2d", "select2d"] }});
    }}
    function renderStats(group) {{
      const selection = ensureSelection(group);
      const body = document.getElementById("stats-body");
      body.innerHTML = "";
      group.plot_channels.filter((channel) => selection.has(channel.name)).forEach((channel) => {{
        const row = document.createElement("tr");
        row.innerHTML = `
          <td><strong>${{channel.name}}</strong></td>
          <td>${{channel.unit || "n/a"}}</td>
          <td>${{channel.stats.sample_count}}</td>
          <td>${{channel.stats.mean ?? "n/a"}}</td>
          <td>${{channel.stats.min ?? "n/a"}}</td>
          <td>${{channel.stats.max ?? "n/a"}}</td>
          <td>${{channel.stats.x_label}}</td>
        `;
        body.appendChild(row);
      }});
      if (!body.children.length) {{
        const row = document.createElement("tr");
        row.innerHTML = `<td colspan="7" class="footnote">선택된 채널이 없습니다.</td>`;
        body.appendChild(row);
      }}
      const scalarGrid = document.getElementById("scalar-grid");
      scalarGrid.innerHTML = "";
      group.scalar_channels.forEach((channel) => {{
        const card = document.createElement("div");
        card.className = "scalar-card";
        card.innerHTML = `
          <strong>${{channel.name}}</strong>
          <div class="footnote">${{channel.unit || "unit n/a"}} · ${{channel.sample_count}} samples</div>
          <code>${{channel.values.join("\\n")}}</code>
        `;
        scalarGrid.appendChild(card);
      }});
    }}
    function renderExcelIntrusionResults() {{
      const chart = excelIntrusion();
      const pdfOverlay = chart.pdf_overlay || {{ available: false, matched_count: 0, rows: [] }};
      const subtitle = document.getElementById("excel-intrusion-subtitle");
      const chips = document.getElementById("excel-intrusion-chips");
      const note = document.getElementById("excel-intrusion-note");
      const body = document.getElementById("excel-intrusion-body");
      chips.innerHTML = "";
      if (!chart.available) {{
        subtitle.textContent = "No Excel intrusion chart data is available for this case.";
        note.textContent = "";
        body.innerHTML = '<tr><td colspan="6" class="footnote">No Excel intrusion chart data is available.</td></tr>';
        Plotly.react("excel-intrusion-plot", [], {{
          margin: {{ l: 24, r: 24, t: 18, b: 24 }},
          paper_bgcolor: "#fffaf2",
          plot_bgcolor: "#fffdf8",
          font: {{ family: "IBM Plex Sans, sans-serif", color: "#1d1814" }},
          annotations: [{{ text: "No Excel intrusion chart", x: .5, y: .5, xref: "paper", yref: "paper", showarrow: false, font: {{ size: 18, color: "#6a6054" }} }}],
          xaxis: {{ visible: false }},
          yaxis: {{ visible: false }},
        }}, {{ responsive: true, displaylogo: false }});
        return;
      }}
      subtitle.textContent = `${{chart.title}} | ${{chart.source_filename}} | sheet ${{chart.source_sheet}}`;
      [
        `Locations: ${{chart.rows.length}}`,
        `Axis max: ${{chart.axis_max_cm}} cm`,
        `Workbook: ${{chart.source_filename}}`,
        pdfOverlay.available ? `PDF overlay: ${{pdfOverlay.matched_count}} / ${{pdfOverlay.row_count}}` : "",
      ].forEach((text) => {{
        if (!text) return;
        const chip = document.createElement("div");
        chip.className = "chip";
        chip.textContent = text;
        chips.appendChild(chip);
      }});
      const xLabels = chart.rows.map((row) => row.label_html);
      const good = chart.rows.map((row) => row.good_cm);
      const acceptable = chart.rows.map((row) => row.acceptable_upper_cm);
      const marginal = chart.rows.map((row) => row.marginal_upper_cm);
      const poor = chart.rows.map(() => chart.axis_max_cm);
      const measured = chart.rows.map((row) => row.measured_cm);
      const pdfMeasured = chart.rows.map((_, index) => pdfOverlay.rows[index]?.pdf_measured_cm ?? null);
      const tickVals = Array.from({{ length: Math.floor(chart.axis_max_cm / 5) + 1 }}, (_, index) => index * 5);
      const primaryBand = chart.rows[0] || {{ good_cm: 15, acceptable_upper_cm: 22.5, marginal_upper_cm: 30 }};
      const traces = [
        {{
          type: "scatter",
          mode: "lines",
          x: xLabels,
          y: good,
          fill: "tozeroy",
          fillcolor: "rgba(85, 240, 40, 0.95)",
          line: {{ color: "#101010", width: 1.2 }},
          hovertemplate: "<b>%{{x}}</b><br>good <= %{{y:.1f}} cm<extra></extra>",
          name: "Good",
        }},
        {{
          type: "scatter",
          mode: "lines",
          x: xLabels,
          y: acceptable,
          fill: "tonexty",
          fillcolor: "rgba(240, 228, 70, 0.92)",
          line: {{ color: "#101010", width: 1.2 }},
          hovertemplate: "<b>%{{x}}</b><br>acceptable <= %{{y:.1f}} cm<extra></extra>",
          name: "Acceptable",
        }},
        {{
          type: "scatter",
          mode: "lines",
          x: xLabels,
          y: marginal,
          fill: "tonexty",
          fillcolor: "rgba(244, 148, 45, 0.9)",
          line: {{ color: "#101010", width: 1.2 }},
          hovertemplate: "<b>%{{x}}</b><br>marginal <= %{{y:.1f}} cm<extra></extra>",
          name: "Marginal",
        }},
        {{
          type: "scatter",
          mode: "lines",
          x: xLabels,
          y: poor,
          fill: "tonexty",
          fillcolor: "rgba(226, 102, 102, 0.82)",
          line: {{ color: "#101010", width: 1.2 }},
          hovertemplate: "<b>%{{x}}</b><br>poor > %{{customdata:.1f}} cm<extra></extra>",
          customdata: marginal,
          name: "Poor",
        }},
        {{
          type: "scatter",
          mode: "lines+markers",
          x: xLabels,
          y: measured,
          line: {{ color: "#2447d6", width: 3 }},
          marker: {{ color: "#2447d6", size: 8 }},
          hovertemplate: "<b>%{{x}}</b><br>measured = %{{y:.1f}} cm<extra></extra>",
          name: "Measured",
        }},
      ];
      if (pdfOverlay.available) {{
        traces.push({{
          type: "scatter",
          mode: "lines+markers",
          x: xLabels,
          y: pdfMeasured,
          line: {{ color: "#111111", width: 2.2, dash: "dash" }},
          marker: {{ color: "#111111", size: 7, symbol: "diamond" }},
          hovertemplate: "<b>%{{x}}</b><br>pdf extracted = %{{y:.1f}} cm<extra></extra>",
          name: "PDF extracted",
        }});
      }}
      Plotly.react("excel-intrusion-plot", traces, {{
        margin: {{ l: 84, r: 64, t: 20, b: 110 }},
        paper_bgcolor: "#fffaf2",
        plot_bgcolor: "#d9d9d9",
        font: {{ family: "IBM Plex Sans, sans-serif", color: "#111111" }},
        showlegend: false,
        xaxis: {{ type: "category", automargin: true, tickfont: {{ size: 11, color: "#111111" }}, showgrid: false, linecolor: "#101010", linewidth: 1, mirror: true, ticks: "outside", fixedrange: true }},
        yaxis: {{ title: "Measured Intrusion (cm)", range: [0, chart.axis_max_cm], tickvals: tickVals, showgrid: false, zeroline: false, linecolor: "#101010", linewidth: 1, mirror: true, ticks: "outside", fixedrange: true }},
        yaxis2: {{ overlaying: "y", side: "right", range: [0, chart.axis_max_cm], tickvals: tickVals, showgrid: false, zeroline: false, linecolor: "#101010", linewidth: 1, ticks: "outside", fixedrange: true }},
        annotations: [
          {{ text: "GOOD", x: .28, y: primaryBand.good_cm * 0.67, xref: "paper", yref: "y", showarrow: false, font: {{ family: "Space Grotesk, sans-serif", size: 18, color: "#0b2508" }} }},
          {{ text: "ACCEPTABLE", x: .31, y: (primaryBand.good_cm + primaryBand.acceptable_upper_cm) / 2, xref: "paper", yref: "y", showarrow: false, font: {{ family: "Space Grotesk, sans-serif", size: 17, color: "#3f3a00" }} }},
          {{ text: "MARGINAL", x: .31, y: (primaryBand.acceptable_upper_cm + primaryBand.marginal_upper_cm) / 2, xref: "paper", yref: "y", showarrow: false, font: {{ family: "Space Grotesk, sans-serif", size: 17, color: "#3d1600" }} }},
          {{ text: "POOR", x: .31, y: (primaryBand.marginal_upper_cm + chart.axis_max_cm) / 2, xref: "paper", yref: "y", showarrow: false, font: {{ family: "Space Grotesk, sans-serif", size: 18, color: "#2c0000" }} }},
        ],
      }}, {{ responsive: true, displaylogo: false, modeBarButtonsToRemove: ["lasso2d", "select2d"] }});
      note.textContent = pdfOverlay.available
        ? `Source: ${{chart.source_filename}} | Blue line shows Excel measured intrusion. Black dashed line shows values parsed from the PDF intrusion table.`
        : `Source: ${{chart.source_filename}} | Blue line shows measured intrusion against Excel rating thresholds.`;
      body.innerHTML = "";
      chart.rows.forEach((row) => {{
        const overlayRow = pdfOverlay.rows.find((entry) => entry.label === row.label) || {{ pdf_measured_cm: null, pdf_rating: "n/a" }};
        const tr = document.createElement("tr");
        const measuredText = row.measured_cm === null || row.measured_cm === undefined ? "n/a" : row.measured_cm.toFixed(1);
        const pdfText = overlayRow.pdf_measured_cm === null || overlayRow.pdf_measured_cm === undefined ? "n/a" : overlayRow.pdf_measured_cm.toFixed(1);
        const deltaText = row.measured_cm === null || row.measured_cm === undefined || overlayRow.pdf_measured_cm === null || overlayRow.pdf_measured_cm === undefined
          ? "n/a"
          : (overlayRow.pdf_measured_cm - row.measured_cm).toFixed(1);
        tr.innerHTML = `
          <td><strong>${{esc(row.label)}}</strong></td>
          <td>${{esc(measuredText)}}</td>
          <td>${{esc(pdfText)}}</td>
          <td>${{esc(deltaText)}}</td>
          <td>${{row.good_cm.toFixed(1)}}</td>
          <td>${{row.acceptable_upper_cm.toFixed(1)}}</td>
          <td>${{row.marginal_upper_cm.toFixed(1)}}</td>
          <td><span class="${{ratingClass(overlayRow.pdf_rating)}}">${{esc(overlayRow.pdf_rating)}}</span></td>
        `;
        body.appendChild(tr);
      }});
    }}
    function pdfTableHasSeat(table) {{
      return Boolean(table?.rows?.some((row) => row.seat_position));
    }}
    function pdfHeadersForTable(table) {{
      const seatHeader = pdfTableHasSeat(table) ? ["Seat"] : [];
      if (["head_injury","neck_injury","chest_injury"].includes(table.table_type)) return [...seatHeader, "Measure", "Threshold", "Result", "Time (ms)"];
      if (table.table_type === "leg_foot_injury") return [...seatHeader, "Section", "Measure", "Threshold", "Left", "Left time", "Right", "Right time"];
      if (table.table_type === "thigh_hip_injury") return [...seatHeader, "Section", "Measure", "Left", "Left time", "Right", "Right time"];
      if (table.table_type === "intrusion") return ["Location", "Longitudinal", "Lateral", "Vertical", "Resultant"];
      if (table.table_type === "dummy_clearance") return ["Code", "Location", "Measure", "Unit"];
      if (table.table_type === "restraint_kinematics") return ["Event", "Time (ms)"];
      return ["Measure", "Value"];
    }}
    function pdfCellsForRow(table, row) {{
      const seatCell = pdfTableHasSeat(table) ? [row.seat_position] : [];
      if (["head_injury","neck_injury","chest_injury"].includes(table.table_type)) return [...seatCell, row.label, row.threshold_text, row.result_text, row.time_text];
      if (table.table_type === "leg_foot_injury") return [...seatCell, row.section_name, row.label, row.threshold_text, row.left_text, row.left_time_text, row.right_text, row.right_time_text];
      if (table.table_type === "thigh_hip_injury") return [...seatCell, row.section_name, row.label, row.left_text, row.left_time_text, row.right_text, row.right_time_text];
      if (table.table_type === "intrusion") return [row.label, row.longitudinal_text, row.lateral_text, row.vertical_text, row.resultant_text];
      if (table.table_type === "dummy_clearance") return [row.code, row.label, row.measure_text, row.unit];
      if (table.table_type === "restraint_kinematics") return [row.label, row.time_text];
      return [row.label || row.normalized_label || row.measure_text || "n/a", row.result_text || row.measure_text || row.left_text || row.right_text || row.time_text || "n/a"];
    }}
    function renderPdfSummary() {{
      const pdf = pdfResults();
      const summaryKpis = document.getElementById("pdf-summary-kpis");
      const summaryNote = document.getElementById("pdf-summary-note");
      const body = document.getElementById("pdf-type-body");
      summaryKpis.innerHTML = "";
      [
        `PDFs: ${{pdf.document_count}}`,
        `Parsed tables: ${{pdf.table_count}}`,
        `Parsed rows: ${{pdf.row_count}}`,
        `Review rows: ${{pdf.review_row_count}}`,
      ].forEach((text) => {{
        const chip = document.createElement("div");
        chip.className = "chip";
        chip.textContent = text;
        summaryKpis.appendChild(chip);
      }});
      if (!pdf.available) {{
        summaryNote.textContent = "No parsed PDF result tables are available for this case.";
        body.innerHTML = '<tr><td colspan="4" class="footnote">No parsed PDF result tables are available.</td></tr>';
        Plotly.react("pdf-type-plot", [], {{
          margin: {{ l: 24, r: 18, t: 10, b: 24 }},
          paper_bgcolor: "#fffaf2",
          plot_bgcolor: "#fffdf8",
          font: {{ family: "IBM Plex Sans, sans-serif", color: "#1d1814" }},
          annotations: [{{ text: "No parsed PDF tables", x: .5, y: .5, xref: "paper", yref: "paper", showarrow: false, font: {{ size: 18, color: "#6a6054" }} }}],
          xaxis: {{ visible: false }},
          yaxis: {{ visible: false }},
        }}, {{ responsive: true, displaylogo: false }});
        return;
      }}
      const rows = [...pdf.type_summary].sort((left, right) => Number(right.row_count) - Number(left.row_count));
      summaryNote.textContent = `Showing ${{rows.length}} parsed result types from ${{pdf.document_count}} PDF document(s).`;
      body.innerHTML = "";
      rows.forEach((row) => {{
        const tr = document.createElement("tr");
        tr.innerHTML = `
          <td><strong>${{esc(prettyPdfType(row.table_type))}}</strong></td>
          <td>${{row.table_count}}</td>
          <td>${{row.row_count}}</td>
          <td>${{row.review_row_count || 0}}</td>
        `;
        body.appendChild(tr);
      }});
      Plotly.react("pdf-type-plot", [
        {{
          type: "bar",
          orientation: "h",
          x: rows.map((row) => row.row_count),
          y: rows.map((row) => prettyPdfType(row.table_type)),
          name: "Rows",
          marker: {{ color: "#0f7173" }},
          hovertemplate: "<b>%{{y}}</b><br>rows=%{{x}}<extra></extra>",
        }},
        {{
          type: "bar",
          orientation: "h",
          x: rows.map((row) => row.review_row_count || 0),
          y: rows.map((row) => prettyPdfType(row.table_type)),
          name: "Review rows",
          marker: {{ color: "#bc4b32" }},
          hovertemplate: "<b>%{{y}}</b><br>review rows=%{{x}}<extra></extra>",
        }},
      ], {{
        barmode: "group",
        margin: {{ l: 210, r: 18, t: 10, b: 48 }},
        paper_bgcolor: "#fffaf2",
        plot_bgcolor: "#fffdf8",
        font: {{ family: "IBM Plex Sans, sans-serif", color: "#1d1814" }},
        xaxis: {{ title: "Parsed rows", gridcolor: "rgba(0,0,0,.09)" }},
        yaxis: {{ automargin: true }},
        legend: {{ orientation: "h", y: 1.14 }},
      }}, {{ responsive: true, displaylogo: false }});
    }}
    function renderPdfTableList() {{
      const list = document.getElementById("pdf-table-list");
      const pdf = pdfResults();
      list.innerHTML = "";
      if (!pdf.available) {{
        list.innerHTML = '<div class="pdf-empty">No parsed PDF table is available.</div>';
        return;
      }}
      pdf.tables.forEach((table) => {{
        const button = document.createElement("button");
        button.type = "button";
        button.className = "pdf-table-btn" + (table.pdf_result_table_id === state.pdfTableId ? " active" : "");
        button.innerHTML = `
          <strong>${{esc(prettyPdfType(table.table_type))}}</strong>
          <div class="muted">${{esc(table.table_title || table.table_ref || "Untitled table")}}</div>
          <div class="muted">Page ${{displayValue(table.page_number)}} | ${{table.row_count}} rows${{table.review_row_count ? ` | review ${{table.review_row_count}}` : ""}}</div>
        `;
        button.onclick = () => {{
          state.pdfTableId = table.pdf_result_table_id;
          renderPdfTableList();
          renderPdfTableDetail();
        }};
        list.appendChild(button);
      }});
    }}
    function renderPdfTableDetail() {{
      const title = document.getElementById("pdf-table-title");
      const subtitle = document.getElementById("pdf-table-subtitle");
      const chips = document.getElementById("pdf-table-chips");
      const headRow = document.getElementById("pdf-table-head-row");
      const body = document.getElementById("pdf-table-body");
      const note = document.getElementById("pdf-table-note");
      const table = selectedPdfTable();
      if (!table) {{
        title.textContent = "Parsed Table Detail";
        subtitle.textContent = "No parsed PDF table is available for this case.";
        chips.innerHTML = "";
        headRow.innerHTML = "";
        body.innerHTML = '<tr><td class="footnote">No parsed PDF rows are available.</td></tr>';
        note.textContent = "";
        return;
      }}
      title.textContent = table.table_title || prettyPdfType(table.table_type);
      subtitle.textContent = `${{table.filename}} | Page ${{displayValue(table.page_number)}} | ${{table.family_label || "n/a"}}`;
      chips.innerHTML = "";
      [
        prettyPdfType(table.table_type),
        `Role: ${{displayValue(table.pdf_role)}}`,
        `Rows: ${{table.row_count}}`,
        `Review rows: ${{table.review_row_count || 0}}`,
        table.table_ref ? `Ref: ${{table.table_ref}}` : "",
      ].filter(Boolean).forEach((text) => {{
        const chip = document.createElement("div");
        chip.className = "chip";
        chip.textContent = text;
        chips.appendChild(chip);
      }});
      const headers = pdfHeadersForTable(table);
      headRow.innerHTML = headers.map((header) => `<th>${{esc(header)}}</th>`).join("");
      body.innerHTML = "";
      table.rows.forEach((row) => {{
        const tr = document.createElement("tr");
        tr.innerHTML = pdfCellsForRow(table, row).map((cell) => `<td>${{esc(displayValue(cell))}}</td>`).join("");
        body.appendChild(tr);
      }});
      if (!table.rows.length) {{
        body.innerHTML = `<tr><td colspan="${{headers.length}}" class="footnote">No parsed rows were captured for this table.</td></tr>`;
      }}
      note.textContent = `${{table.rows.length}} parsed rows are displayed for the selected PDF table.`;
    }}
    function renderPdfResults() {{
      renderPdfSummary();
      renderPdfTableList();
      renderPdfTableDetail();
    }}
    function render() {{
      const group = currentGroup();
      ensureSelection(group);
      renderGroupButtons();
      renderHeader(group);
      renderChannelList(group);
      renderPlot(group);
      renderStats(group);
    }}
    document.getElementById("prev-group").onclick = () => {{ state.groupIndex = (state.groupIndex - 1 + dashboard.groups.length) % dashboard.groups.length; state.filter = ""; document.getElementById("channel-filter").value = ""; render(); }};
    document.getElementById("next-group").onclick = () => {{ state.groupIndex = (state.groupIndex + 1) % dashboard.groups.length; state.filter = ""; document.getElementById("channel-filter").value = ""; render(); }};
    document.getElementById("channel-filter").oninput = (event) => {{ state.filter = event.target.value; renderChannelList(currentGroup()); }};
    document.getElementById("select-default").onclick = () => {{ const group = currentGroup(); state.selections[group.name] = new Set(defaultSelection(group)); render(); }};
    document.getElementById("select-visible").onclick = () => {{ const group = currentGroup(); state.selections[group.name] = new Set(filteredChannels(group).map((channel) => channel.name)); render(); }};
    document.getElementById("clear-selection").onclick = () => {{ const group = currentGroup(); state.selections[group.name] = new Set(); render(); }};
    window.addEventListener("keydown", (event) => {{
      if (event.key === "ArrowLeft") document.getElementById("prev-group").click();
      if (event.key === "ArrowRight") document.getElementById("next-group").click();
    }});
    render();
    renderExcelIntrusionResults();
    renderPdfResults();
  </script>
</body>
</html>"""


def main() -> None:
    args = parse_args()
    input_root = resolve_repo_path(args.input_root, INPUT_ROOT)
    output_root = resolve_repo_path(args.output_root, OUTPUT_ROOT)
    plots_root = resolve_repo_path(args.plots_root, PLOTS_ROOT)
    research_db_path = resolve_repo_path(args.research_db, RESEARCH_DB)
    case_root = find_case_root(input_root, args.filegroup_id, args.test_code)
    dashboard_dir = output_root / case_root.name
    dashboard_dir.mkdir(parents=True, exist_ok=True)
    data = build_dashboard_data(case_root, dashboard_dir, plots_root, output_root, research_db_path)
    output_path = dashboard_dir / "index.html"
    output_path.write_text(dashboard_html(data), encoding="utf-8")
    print(json.dumps({"dashboard_html": str(output_path), "dashboard_groups": data["summary"]["dashboard_group_count"], "tdms_groups": data["summary"]["tdms_group_count"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
