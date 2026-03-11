from __future__ import annotations

import argparse
import json
import math
import re
import sqlite3
import time
import warnings
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from scripts.core.catalog import excel_catalog_schema
import openpyxl
import pandas as pd
import xlrd
from openpyxl.chartsheet.chartsheet import Chartsheet


PARSER_VERSION = "excel-pipeline:v1"
REPO_ROOT = Path(__file__).resolve().parents[3]
SUMMARY_CSV = REPO_ROOT / "output/small_overlap/tables/excel_etl_run_summary.csv"
EXCEL_CATALOG_SQL = """
CREATE INDEX IF NOT EXISTS idx_excel_workbooks_type_status ON excel_workbooks(workbook_type, extraction_status);
CREATE INDEX IF NOT EXISTS idx_extracted_metrics_source_asset ON extracted_metrics(source_type, asset_id);
CREATE INDEX IF NOT EXISTS idx_extracted_metrics_source_namespace_metric ON extracted_metrics(source_type, namespace, metric_name);

DROP VIEW IF EXISTS excel_metric_summary;
DROP VIEW IF EXISTS excel_metric_catalog;
DROP VIEW IF EXISTS excel_workbook_inventory;

CREATE VIEW excel_workbook_inventory AS
WITH sheet_summary AS (
  SELECT excel_workbook_id,
         COUNT(*) AS sheet_count,
         COALESCE(SUM(COALESCE(row_count, 0)), 0) AS total_sheet_rows,
         MAX(COALESCE(column_count, 0)) AS max_column_count
    FROM excel_sheets
   GROUP BY excel_workbook_id
),
metric_summary AS (
  SELECT asset_id,
         COUNT(*) AS metric_count,
         COUNT(DISTINCT namespace) AS namespace_count,
         COUNT(DISTINCT namespace || '|' || metric_name || '|' || COALESCE(metric_unit, '')) AS distinct_metric_count,
         ROUND(AVG(confidence), 3) AS avg_confidence
    FROM extracted_metrics
   WHERE source_type = 'excel_workbook'
   GROUP BY asset_id
),
namespace_summary AS (
  SELECT asset_id,
         GROUP_CONCAT(namespace_entry, ', ') AS namespace_counts
    FROM (
      SELECT asset_id,
             namespace || ':' || namespace_metric_count AS namespace_entry
        FROM (
          SELECT asset_id, namespace, COUNT(*) AS namespace_metric_count
            FROM extracted_metrics
           WHERE source_type = 'excel_workbook'
           GROUP BY asset_id, namespace
        )
       ORDER BY asset_id, namespace
    )
   GROUP BY asset_id
)
SELECT ew.excel_workbook_id,
       ew.asset_id,
       ew.filegroup_id,
       fg.test_code,
       fg.title AS filegroup_title,
       fg.tested_on,
       tt.test_type_label,
       v.vehicle_year,
       v.vehicle_make_model,
       ew.workbook_type,
       ew.extraction_status,
       ew.notes,
       a.filename,
       a.local_path,
       a.relative_path,
       a.folder_path,
       COALESCE(ss.sheet_count, 0) AS sheet_count,
       COALESCE(ss.total_sheet_rows, 0) AS total_sheet_rows,
       COALESCE(ss.max_column_count, 0) AS max_column_count,
       COALESCE(ms.metric_count, 0) AS metric_count,
       COALESCE(ms.namespace_count, 0) AS namespace_count,
       COALESCE(ms.distinct_metric_count, 0) AS distinct_metric_count,
       ms.avg_confidence,
       ns.namespace_counts
  FROM excel_workbooks ew
  JOIN assets a ON a.asset_id = ew.asset_id
  JOIN filegroups fg ON fg.filegroup_id = ew.filegroup_id
  LEFT JOIN test_types tt ON tt.test_type_code = fg.test_type_code
  LEFT JOIN vehicles v ON v.vehicle_id = fg.vehicle_id
  LEFT JOIN sheet_summary ss ON ss.excel_workbook_id = ew.excel_workbook_id
  LEFT JOIN metric_summary ms ON ms.asset_id = ew.asset_id
  LEFT JOIN namespace_summary ns ON ns.asset_id = ew.asset_id;

CREATE VIEW excel_metric_catalog AS
SELECT em.extracted_metric_id,
       ew.excel_workbook_id,
       ew.asset_id,
       ew.filegroup_id,
       fg.test_code,
       fg.title AS filegroup_title,
       fg.tested_on,
       tt.test_type_label,
       v.vehicle_year,
       v.vehicle_make_model,
       ew.workbook_type,
       ew.extraction_status,
       a.filename,
       a.local_path,
       a.relative_path,
       a.folder_path,
       CASE
         WHEN em.source_locator LIKE 'sheet:%' THEN
           CASE
             WHEN instr(substr(em.source_locator, 7), '|') > 0 THEN substr(substr(em.source_locator, 7), 1, instr(substr(em.source_locator, 7), '|') - 1)
             ELSE substr(em.source_locator, 7)
           END
         ELSE ''
       END AS sheet_name,
       em.source_locator,
       em.namespace,
       em.metric_name,
       em.metric_value_text,
       em.metric_value_number,
       em.metric_unit,
       em.confidence,
       em.extraction_method
  FROM extracted_metrics em
  JOIN excel_workbooks ew ON ew.asset_id = em.asset_id
  JOIN assets a ON a.asset_id = ew.asset_id
  JOIN filegroups fg ON fg.filegroup_id = ew.filegroup_id
  LEFT JOIN test_types tt ON tt.test_type_code = fg.test_type_code
  LEFT JOIN vehicles v ON v.vehicle_id = fg.vehicle_id
 WHERE em.source_type = 'excel_workbook';

CREATE VIEW excel_metric_summary AS
WITH metric_base AS (
  SELECT excel_workbook_id,
         filegroup_id,
         test_code,
         workbook_type,
         namespace,
         metric_name,
         COALESCE(metric_unit, '') AS metric_unit,
         metric_value_number,
         confidence
    FROM excel_metric_catalog
),
sample_codes AS (
  SELECT workbook_type,
         namespace,
         metric_name,
         metric_unit,
         GROUP_CONCAT(test_code, ', ') AS sample_test_codes
    FROM (
      SELECT DISTINCT workbook_type, namespace, metric_name, metric_unit, test_code
        FROM metric_base
       ORDER BY workbook_type, namespace, metric_name, metric_unit, test_code
    )
   GROUP BY workbook_type, namespace, metric_name, metric_unit
)
SELECT mb.workbook_type,
       mb.namespace,
       mb.metric_name,
       mb.metric_unit,
       COUNT(*) AS metric_count,
       COUNT(DISTINCT mb.excel_workbook_id) AS workbook_count,
       COUNT(DISTINCT mb.filegroup_id) AS filegroup_count,
       COUNT(mb.metric_value_number) AS numeric_value_count,
       ROUND(AVG(mb.confidence), 3) AS avg_confidence,
       ROUND(MIN(mb.metric_value_number), 3) AS min_value_number,
       ROUND(MAX(mb.metric_value_number), 3) AS max_value_number,
       sc.sample_test_codes
  FROM metric_base mb
  LEFT JOIN sample_codes sc
    ON sc.workbook_type = mb.workbook_type
   AND sc.namespace = mb.namespace
   AND sc.metric_name = mb.metric_name
   AND sc.metric_unit = mb.metric_unit
 GROUP BY mb.workbook_type, mb.namespace, mb.metric_name, mb.metric_unit, sc.sample_test_codes;
"""


class InvalidWorkbookArtifactError(Exception):
    pass


@dataclass
class ExcelJob:
    excel_workbook_id: int
    asset_id: int
    filegroup_id: int
    test_code: str
    workbook_type: str
    filename: str
    local_path: str
    extraction_status: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract workbook inventory and first-pass metrics into the research DB.")
    parser.add_argument("--db", default="data/research/research.sqlite")
    parser.add_argument("--excel-workbook-id", type=int, default=None)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--commit-interval", type=int, default=25, help="Commit every N processed workbooks.")
    parser.add_argument("--all", action="store_true", help="Process all workbook rows instead of only pending/error.")
    parser.add_argument(
        "--include-done",
        action="store_true",
        help="Include already-done workbook rows when --all is used.",
    )
    return parser.parse_args()


def ensure_excel_catalog_schema(connection: sqlite3.Connection) -> None:
    excel_catalog_schema.ensure_excel_catalog_schema(connection)


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def resolve_repo_path(value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else REPO_ROOT / path


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def looks_like_html_artifact(path: Path) -> bool:
    try:
        snippet = path.read_bytes()[:4096].decode("utf-8", errors="ignore").lower()
    except OSError:
        return False
    markers = ("<!doctype html", "<html", "you are not logged in", "<title>\r\n\tiihs techdata", "<title>\n\tiihs techdata", "<title>iihs techdata")
    return any(marker in snippet for marker in markers)


def slugify(value: str) -> str:
    text = normalize_text(value).lower()
    text = text.replace("%", " percent ")
    text = text.replace("/", " ")
    text = re.sub(r"[\(\)\[\]\{\}]", " ", text)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace(",", "")
    text = re.sub(r"[^0-9.\-+]", "", text)
    if not text or text in {"-", "+", ".", "-.", "+."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_percentish(value: Any) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    if text.endswith("%"):
        number = parse_float(text[:-1])
        return None if number is None else number / 100.0
    return parse_float(text)


def metric_row(
    job: ExcelJob,
    source_type: str,
    source_locator: str,
    namespace: str,
    metric_name: str,
    metric_value_text: str | None,
    metric_value_number: float | None,
    metric_unit: str | None,
    confidence: float,
    extraction_method: str,
) -> tuple[Any, ...]:
    return (
        job.filegroup_id,
        job.asset_id,
        source_type,
        source_locator,
        namespace,
        metric_name,
        metric_value_text,
        metric_value_number,
        metric_unit,
        confidence,
        extraction_method,
    )


def read_excel_table(path: Path, sheet_name: str, extension: str, nrows: int | None = None) -> pd.DataFrame:
    engine = "xlrd" if extension == ".xls" else "openpyxl"
    return pd.read_excel(path, sheet_name=sheet_name, header=None, engine=engine, nrows=nrows)


def open_workbook_for_inventory(path: Path, extension: str):
    if extension == ".xls":
        return xlrd.open_workbook(path, on_demand=True)
    return openpyxl.load_workbook(path, read_only=True, data_only=False)


def close_workbook(workbook: Any, extension: str) -> None:
    if extension == ".xls":
        release = getattr(workbook, "release_resources", None)
        if callable(release):
            release()
        return
    close = getattr(workbook, "close", None)
    if callable(close):
        close()


def workbook_sheet_infos(path: Path, extension: str) -> tuple[list[dict[str, Any]], dict[str, int]]:
    workbook = open_workbook_for_inventory(path, extension)
    sheets: list[dict[str, Any]] = []
    worksheet_count = 0
    chartsheet_count = 0
    try:
        if extension == ".xls":
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                worksheet_count += 1
                sheets.append(
                    {
                        "sheet_name": sheet_name,
                        "row_count": int(sheet.nrows),
                        "column_count": int(sheet.ncols),
                        "sheet_kind": "worksheet",
                    }
                )
        else:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if isinstance(sheet, Chartsheet):
                    chartsheet_count += 1
                    sheets.append({"sheet_name": sheet_name, "row_count": 0, "column_count": 0, "sheet_kind": "chartsheet"})
                    continue
                worksheet_count += 1
                sheets.append(
                    {
                        "sheet_name": sheet_name,
                        "row_count": int(sheet.max_row or 0),
                        "column_count": int(sheet.max_column or 0),
                        "sheet_kind": "worksheet",
                    }
                )
    finally:
        close_workbook(workbook, extension)
    return sheets, {"worksheet_count": worksheet_count, "chartsheet_count": chartsheet_count}


def extract_summary_metrics(job: ExcelJob, path: Path, extension: str, sheet_infos: list[dict[str, Any]]) -> list[tuple[Any, ...]]:
    metrics: list[tuple[Any, ...]] = []
    summary_sheet = next(
        (
            row["sheet_name"]
            for row in sheet_infos
            if "summary" in slugify(row["sheet_name"])
            and row["sheet_kind"] == "worksheet"
        ),
        None,
    )
    ratings_sheet = next(
        (
            row["sheet_name"]
            for row in sheet_infos
            if slugify(row["sheet_name"]) == "ratings_chart_data"
            and row["sheet_kind"] == "worksheet"
        ),
        None,
    )
    if summary_sheet:
        df = read_excel_table(path, summary_sheet, extension, nrows=80).fillna("")
        row0 = [normalize_text(value) for value in df.iloc[0].tolist()]
        row1 = [normalize_text(value) for value in df.iloc[1].tolist()]
        row2 = [normalize_text(value) for value in df.iloc[2].tolist()]
        if row0:
            metrics.append(metric_row(job, "excel_workbook", f"sheet:{summary_sheet}", "excel_summary", "summary_title", row0[0], None, None, 0.9, "sheet_header"))
        if len(row0) >= 6 and row0[5]:
            metrics.append(metric_row(job, "excel_workbook", f"sheet:{summary_sheet}", "excel_summary", "test_date", row0[5], None, None, 0.9, "sheet_header"))
        if row1 and row1[1]:
            metrics.append(metric_row(job, "excel_workbook", f"sheet:{summary_sheet}", "excel_summary", "test_description", row1[1], None, None, 0.85, "sheet_header"))
        if row2 and row2[0]:
            impact_speed = parse_float(row2[1] if len(row2) > 1 else "")
            metrics.append(metric_row(job, "excel_workbook", f"sheet:{summary_sheet}", "excel_summary", "actual_test_impact_speed_kmh", row2[1] if len(row2) > 1 else None, impact_speed, "km/h", 0.92, "tabular_summary"))

        current_section = ""
        for idx in range(3, len(df)):
            values = [normalize_text(value) for value in df.iloc[idx].tolist()]
            section_candidate = values[0]
            label = values[1] if len(values) > 1 else ""
            if section_candidate and label:
                current_section = slugify(section_candidate)
            if not label or "reference value" in label.lower():
                continue
            base = slugify(f"{current_section}_{label}" if current_section else label)
            if not base:
                continue
            for suffix, position in (("reference", 2), ("measured", 3), ("t1", 4), ("t2", 5)):
                if position >= len(values) or not values[position]:
                    continue
                metric_text = values[position]
                metric_number = parse_float(metric_text)
                metrics.append(
                    metric_row(
                        job,
                        "excel_workbook",
                        f"sheet:{summary_sheet}|row:{idx + 1}",
                        "excel_summary",
                        f"{base}_{suffix}",
                        metric_text,
                        metric_number,
                        None,
                        0.88,
                        "tabular_summary",
                    )
                )

    if ratings_sheet:
        df = read_excel_table(path, ratings_sheet, extension, nrows=80).fillna("")
        for idx in range(2, len(df)):
            values = [normalize_text(value) for value in df.iloc[idx].tolist()]
            label = values[1] if len(values) > 1 and values[1] else (values[0] if values else "")
            if not label:
                continue
            base = slugify(label)
            if len(values) > 9 and values[1]:
                pairs = {
                    "as_measured": 2,
                    "normalized_iarv": 3,
                    "iarv": 6,
                    "boundary_ga": 7,
                    "boundary_am": 8,
                    "boundary_mp": 9,
                }
            else:
                pairs = {
                    "as_measured": 1,
                    "normalized_iarv": 2,
                    "iarv": 3,
                    "boundary_ga": 4,
                    "boundary_am": 5,
                    "boundary_mp": 6,
                }
            for suffix, position in pairs.items():
                if position >= len(values) or not values[position]:
                    continue
                metric_text = values[position]
                metric_number = parse_float(metric_text)
                metrics.append(
                    metric_row(
                        job,
                        "excel_workbook",
                        f"sheet:{ratings_sheet}|row:{idx + 1}",
                        "excel_ratings",
                        f"{base}_{suffix}",
                        metric_text,
                        metric_number,
                        None,
                        0.9,
                        "ratings_chart_data",
                    )
                )
    return metrics


def extract_das_summary_metrics(job: ExcelJob, path: Path, extension: str, sheet_infos: list[dict[str, Any]]) -> list[tuple[Any, ...]]:
    metrics: list[tuple[Any, ...]] = []
    das_sheet = next(
        (
            row["sheet_name"]
            for row in sheet_infos
            if slugify(row["sheet_name"]) == "das"
            and row["sheet_kind"] == "worksheet"
        ),
        None,
    )
    if not das_sheet:
        return metrics
    df = read_excel_table(path, das_sheet, extension, nrows=80).fillna("")
    header_map = {
        "test setup name": "test_setup_name",
        "test description": "test_description",
        "test id": "test_id",
        "report date": "report_date",
        "recording mode": "recording_mode",
    }
    for idx in range(min(8, len(df))):
        values = [normalize_text(value) for value in df.iloc[idx].tolist()]
        label = values[0].lower() if values else ""
        metric_name = header_map.get(label)
        value = values[2] if len(values) > 2 else ""
        if metric_name and value:
            metrics.append(
                metric_row(
                    job,
                    "excel_workbook",
                    f"sheet:{das_sheet}|row:{idx + 1}",
                    "excel_das_summary",
                    metric_name,
                    value,
                    parse_float(value),
                    None,
                    0.92,
                    "das_summary_header",
                )
            )

    module_count = 0
    active_channel_total = 0
    sample_rates: set[float] = set()
    for idx in range(7, len(df)):
        values = [normalize_text(value) for value in df.iloc[idx].tolist()]
        serial = values[0] if values else ""
        if not serial:
            continue
        module_count += 1
        if len(values) > 8:
            sample_rate = parse_float(values[8])
            if sample_rate is not None:
                sample_rates.add(sample_rate)
        if len(values) > 13 and values[13]:
            match = re.match(r"(\d+)", values[13])
            if match:
                active_channel_total += int(match.group(1))
    if module_count:
        metrics.append(metric_row(job, "excel_workbook", f"sheet:{das_sheet}", "excel_das_summary", "module_count", str(module_count), float(module_count), "modules", 0.9, "das_module_table"))
    if active_channel_total:
        metrics.append(metric_row(job, "excel_workbook", f"sheet:{das_sheet}", "excel_das_summary", "active_channel_total", str(active_channel_total), float(active_channel_total), "channels", 0.9, "das_module_table"))
    if sample_rates:
        sample_rate = min(sample_rates)
        metrics.append(metric_row(job, "excel_workbook", f"sheet:{das_sheet}", "excel_das_summary", "sample_rate_hz", f"{sample_rate}", sample_rate, "Hz", 0.9, "das_module_table"))

    sensor_sheet = next(
        (
            row["sheet_name"]
            for row in sheet_infos
            if slugify(row["sheet_name"]) == "sensor_channels"
            and row["sheet_kind"] == "worksheet"
        ),
        None,
    )
    if sensor_sheet:
        sensor_df = read_excel_table(path, sensor_sheet, extension).fillna("")
        sensor_channel_rows = 0
        for idx in range(len(sensor_df)):
            values = [normalize_text(value) for value in sensor_df.iloc[idx].tolist()]
            if not values or not values[0].isdigit():
                continue
            sensor_channel_rows += 1
        if sensor_channel_rows:
            metrics.append(
                metric_row(
                    job,
                    "excel_workbook",
                    f"sheet:{sensor_sheet}",
                    "excel_das_summary",
                    "sensor_channel_count",
                    str(sensor_channel_rows),
                    float(sensor_channel_rows),
                    "channels",
                    0.92,
                    "sensor_channel_inventory",
                )
            )
    return metrics


def extract_intrusion_metrics(job: ExcelJob, path: Path, extension: str, sheet_infos: list[dict[str, Any]]) -> list[tuple[Any, ...]]:
    metrics: list[tuple[Any, ...]] = []
    candidate = next((row["sheet_name"] for row in sheet_infos if normalize_text(row["sheet_name"]).upper() == job.test_code), None)
    if not candidate:
        candidate = next((row["sheet_name"] for row in sheet_infos if normalize_text(row["sheet_name"]).startswith(job.test_code)), None)
    if not candidate:
        return metrics
    df = read_excel_table(path, candidate, extension).fillna("")
    skip_labels = {
        "description",
        "measurement",
        "target",
        "origin_start",
        "origin_finish",
        "pre_crash_origin",
        "post_crash_origin",
        "start",
        "finish",
    }
    for idx in range(len(df)):
        values = [normalize_text(value) for value in df.iloc[idx].tolist()]
        label = values[1] if len(values) > 1 and values[1] else (values[0] if values else "")
        label_slug = slugify(label)
        if not label or label_slug in skip_labels:
            continue
        numeric = {position: parse_float(value) for position, value in enumerate(values)}
        numeric = {position: value for position, value in numeric.items() if value is not None}
        modern_layout = all(position in numeric for position in (2, 3, 4, 5, 6, 7))
        legacy_layout = all(position in numeric for position in (2, 3, 4, 10, 11, 12))
        if not modern_layout and not legacy_layout:
            continue

        base = label_slug
        if modern_layout:
            pre_positions = (2, 3, 4)
            post_positions = (5, 6, 7)
            diff_positions = (8, 9, 10)
            resultant_position = 11
            longitudinal_position = 12
        else:
            pre_positions = (2, 3, 4)
            post_positions = (10, 11, 12)
            diff_positions = (None, None, None)
            resultant_position = None
            longitudinal_position = None

        axes = ("x", "y", "z")
        diff_values: list[float] = []
        for axis, pre_position, post_position, diff_position in zip(axes, pre_positions, post_positions, diff_positions):
            pre_value = numeric.get(pre_position)
            post_value = numeric.get(post_position)
            if pre_value is None or post_value is None:
                continue
            metrics.append(
                metric_row(
                    job,
                    "excel_workbook",
                    f"sheet:{candidate}|row:{idx + 1}",
                    "excel_intrusion",
                    f"{base}_pre_{axis}_mm",
                    f"{pre_value}",
                    pre_value,
                    "mm",
                    0.9,
                    "intrusion_geometry_table",
                )
            )
            metrics.append(
                metric_row(
                    job,
                    "excel_workbook",
                    f"sheet:{candidate}|row:{idx + 1}",
                    "excel_intrusion",
                    f"{base}_post_{axis}_mm",
                    f"{post_value}",
                    post_value,
                    "mm",
                    0.9,
                    "intrusion_geometry_table",
                )
            )
            diff_value = abs(pre_value - post_value)
            if diff_position is not None and diff_position in numeric:
                diff_value = abs(float(numeric[diff_position]))
            diff_values.append(diff_value)
            metrics.append(
                metric_row(
                    job,
                    "excel_workbook",
                    f"sheet:{candidate}|row:{idx + 1}",
                    "excel_intrusion",
                    f"{base}_diff_{axis}_mm",
                    f"{diff_value}",
                    diff_value,
                    "mm",
                    0.9,
                    "intrusion_geometry_table",
                )
            )
        if diff_values:
            resultant_value = math.sqrt(sum(value * value for value in diff_values))
            if resultant_position is not None and resultant_position in numeric:
                resultant_value = abs(float(numeric[resultant_position]))
            metrics.append(
                metric_row(
                    job,
                    "excel_workbook",
                    f"sheet:{candidate}|row:{idx + 1}",
                    "excel_intrusion",
                    f"{base}_resultant_mm",
                    f"{resultant_value}",
                    resultant_value,
                    "mm",
                    0.9,
                    "intrusion_geometry_table",
                )
            )
        if longitudinal_position is not None and longitudinal_position in numeric:
            longitudinal_value = float(numeric[longitudinal_position])
            metrics.append(
                metric_row(
                    job,
                    "excel_workbook",
                    f"sheet:{candidate}|row:{idx + 1}",
                    "excel_intrusion",
                    f"{base}_longitudinal_x_cm",
                    f"{longitudinal_value}",
                    longitudinal_value,
                    "cm",
                    0.88,
                    "intrusion_geometry_table",
                )
            )
        metrics.append(
            metric_row(
                job,
                "excel_workbook",
                f"sheet:{candidate}|row:{idx + 1}",
                "excel_intrusion",
                f"{base}_label",
                label,
                None,
                None,
                0.95,
                "intrusion_geometry_table",
            )
        )
    return metrics


def extract_environment_metrics(job: ExcelJob, path: Path, extension: str, sheet_infos: list[dict[str, Any]]) -> list[tuple[Any, ...]]:
    metrics: list[tuple[Any, ...]] = []
    candidate = sheet_infos[0]["sheet_name"] if sheet_infos else None
    if not candidate:
        return metrics
    df = pd.read_excel(path, sheet_name=candidate, engine="openpyxl")
    if df.empty:
        return metrics
    temperature = pd.to_numeric(df.get("Temperature(F)"), errors="coerce")
    humidity = df.get("Humidity")
    humidity_ratio = humidity.map(parse_percentish) if humidity is not None else pd.Series(dtype=float)
    date_text = df.get("Date")
    time_text = df.get("Time")
    if date_text is not None and time_text is not None:
        timestamps = pd.to_datetime(date_text.astype(str) + " " + time_text.astype(str), errors="coerce")
    else:
        timestamps = pd.Series(dtype="datetime64[ns]")

    summary_values = {
        "row_count": (str(len(df)), float(len(df)), "rows"),
        "temperature_avg_f": (f"{temperature.mean():.3f}" if temperature.notna().any() else None, float(temperature.mean()) if temperature.notna().any() else None, "degF"),
        "temperature_min_f": (f"{temperature.min():.3f}" if temperature.notna().any() else None, float(temperature.min()) if temperature.notna().any() else None, "degF"),
        "temperature_max_f": (f"{temperature.max():.3f}" if temperature.notna().any() else None, float(temperature.max()) if temperature.notna().any() else None, "degF"),
        "humidity_avg_ratio": (f"{humidity_ratio.mean():.6f}" if humidity_ratio.notna().any() else None, float(humidity_ratio.mean()) if humidity_ratio.notna().any() else None, "ratio"),
    }
    for metric_name, (text_value, number_value, unit) in summary_values.items():
        if text_value is None and number_value is None:
            continue
        metrics.append(metric_row(job, "excel_workbook", f"sheet:{candidate}", "excel_environment", metric_name, text_value, number_value, unit, 0.95, "environment_summary"))
    if timestamps.notna().any():
        metrics.append(metric_row(job, "excel_workbook", f"sheet:{candidate}", "excel_environment", "timestamp_start", str(timestamps.min()), None, None, 0.95, "environment_summary"))
        metrics.append(metric_row(job, "excel_workbook", f"sheet:{candidate}", "excel_environment", "timestamp_end", str(timestamps.max()), None, None, 0.95, "environment_summary"))
    if "Location" in df.columns and df["Location"].notna().any():
        metrics.append(metric_row(job, "excel_workbook", f"sheet:{candidate}", "excel_environment", "location", normalize_text(df["Location"].dropna().iloc[0]), None, None, 0.9, "environment_summary"))
    return metrics


def extract_umtri_metrics(job: ExcelJob, path: Path, extension: str, sheet_infos: list[dict[str, Any]]) -> list[tuple[Any, ...]]:
    metrics: list[tuple[Any, ...]] = []
    candidate = next((row["sheet_name"] for row in sheet_infos if row["sheet_name"] == "SEAT INFORMATION"), None)
    if not candidate:
        return metrics
    df = read_excel_table(path, candidate, extension, nrows=40).fillna("")
    for idx in range(len(df)):
        values = [normalize_text(value) for value in df.iloc[idx].tolist()]
        label = values[1] if len(values) > 1 else ""
        value = values[2] if len(values) > 2 else ""
        if not label or not value:
            continue
        if label in {"Adjustment type", "Adjustment range (mm)"}:
            unit = "mm" if label.endswith("(mm)") else None
            metrics.append(
                metric_row(
                    job,
                    "excel_workbook",
                    f"sheet:{candidate}|row:{idx + 1}",
                    "excel_umtri",
                    slugify(label),
                    value,
                    parse_float(value),
                    unit,
                    0.85,
                    "umtri_seat_information",
                )
            )
    return metrics


def extract_generic_inventory_metrics(job: ExcelJob, sheet_infos: list[dict[str, Any]], counts: dict[str, int]) -> list[tuple[Any, ...]]:
    return [
        metric_row(job, "excel_workbook", "workbook", "excel_inventory", "sheet_count", str(len(sheet_infos)), float(len(sheet_infos)), "sheets", 1.0, "inventory"),
        metric_row(job, "excel_workbook", "workbook", "excel_inventory", "worksheet_count", str(counts["worksheet_count"]), float(counts["worksheet_count"]), "worksheets", 1.0, "inventory"),
        metric_row(job, "excel_workbook", "workbook", "excel_inventory", "chartsheet_count", str(counts["chartsheet_count"]), float(counts["chartsheet_count"]), "chartsheets", 1.0, "inventory"),
    ]


def load_jobs(connection: sqlite3.Connection, args: argparse.Namespace) -> list[ExcelJob]:
    query = """
        SELECT ew.excel_workbook_id,
               ew.asset_id,
               ew.filegroup_id,
               fg.test_code,
               ew.workbook_type,
               a.filename,
               a.local_path,
               ew.extraction_status
          FROM excel_workbooks ew
          JOIN assets a ON a.asset_id = ew.asset_id
          JOIN filegroups fg ON fg.filegroup_id = ew.filegroup_id
    """
    clauses = []
    params: list[Any] = []
    if args.excel_workbook_id:
        clauses.append("ew.excel_workbook_id = ?")
        params.append(args.excel_workbook_id)
    elif not args.all:
        clauses.append("ew.extraction_status IN ('pending', 'error')")
    elif not args.include_done:
        clauses.append("ew.extraction_status <> 'done'")
    if clauses:
        query += " WHERE " + " AND ".join(clauses)
    query += " ORDER BY ew.excel_workbook_id"
    if args.limit:
        query += " LIMIT ?"
        params.append(args.limit)
    rows = connection.execute(query, params).fetchall()
    return [ExcelJob(**dict(row)) for row in rows]


def delete_existing_state(connection: sqlite3.Connection, job: ExcelJob) -> None:
    connection.execute("DELETE FROM excel_sheets WHERE excel_workbook_id = ?", (job.excel_workbook_id,))
    connection.execute(
        "DELETE FROM extracted_metrics WHERE asset_id = ? AND source_type IN ('excel_workbook', 'excel_sheet')",
        (job.asset_id,),
    )


def insert_sheet_rows(connection: sqlite3.Connection, job: ExcelJob, sheet_infos: list[dict[str, Any]]) -> None:
    connection.executemany(
        """
        INSERT INTO excel_sheets (
          excel_workbook_id, asset_id, sheet_name, row_count, column_count, extraction_status
        ) VALUES (?, ?, ?, ?, ?, ?)
        """,
        [
            (
                job.excel_workbook_id,
                job.asset_id,
                row["sheet_name"],
                row["row_count"],
                row["column_count"],
                "done",
            )
            for row in sheet_infos
        ],
    )


def insert_metrics(connection: sqlite3.Connection, metrics: list[tuple[Any, ...]]) -> None:
    if not metrics:
        return
    connection.executemany(
        """
        INSERT INTO extracted_metrics (
          filegroup_id,
          asset_id,
          source_type,
          source_locator,
          namespace,
          metric_name,
          metric_value_text,
          metric_value_number,
          metric_unit,
          confidence,
          extraction_method
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        metrics,
    )


def process_job(connection: sqlite3.Connection, job: ExcelJob) -> dict[str, Any]:
    path = Path(job.local_path)
    extension = path.suffix.lower()
    notes: list[str] = []
    delete_existing_state(connection, job)
    try:
        if not path.exists():
            raise FileNotFoundError(path)
        if looks_like_html_artifact(path):
            raise InvalidWorkbookArtifactError("Workbook file contains login HTML instead of workbook content.")
        sheet_infos, counts = workbook_sheet_infos(path, extension)
        insert_sheet_rows(connection, job, sheet_infos)
        metrics = extract_generic_inventory_metrics(job, sheet_infos, counts)
        if job.workbook_type == "summary":
            metrics.extend(extract_summary_metrics(job, path, extension, sheet_infos))
            metrics.extend(extract_das_summary_metrics(job, path, extension, sheet_infos))
            notes.append("summary metrics extracted")
        elif job.workbook_type == "intrusion":
            metrics.extend(extract_intrusion_metrics(job, path, extension, sheet_infos))
            notes.append("intrusion metrics extracted")
        elif job.workbook_type == "environment":
            metrics.extend(extract_environment_metrics(job, path, extension, sheet_infos))
            notes.append("environment metrics extracted")
        elif job.workbook_type == "umtri":
            metrics.extend(extract_umtri_metrics(job, path, extension, sheet_infos))
            notes.append("umtri seat metrics extracted")
        else:
            notes.append("inventory only")
        insert_metrics(connection, metrics)
        connection.execute(
            "UPDATE excel_workbooks SET extraction_status = ?, notes = ? WHERE excel_workbook_id = ?",
            ("done", f"{PARSER_VERSION}; {', '.join(notes)}; sheets={len(sheet_infos)}; metrics={len(metrics)}", job.excel_workbook_id),
        )
        return {
            "excel_workbook_id": job.excel_workbook_id,
            "filegroup_id": job.filegroup_id,
            "test_code": job.test_code,
            "workbook_type": job.workbook_type,
            "status": "done",
            "sheet_count": len(sheet_infos),
            "metric_count": len(metrics),
            "notes": "; ".join(notes),
        }
    except InvalidWorkbookArtifactError as exc:
        connection.execute(
            "UPDATE excel_workbooks SET extraction_status = ?, notes = ? WHERE excel_workbook_id = ?",
            ("skipped", f"{PARSER_VERSION}; {type(exc).__name__}: {exc}", job.excel_workbook_id),
        )
        return {
            "excel_workbook_id": job.excel_workbook_id,
            "filegroup_id": job.filegroup_id,
            "test_code": job.test_code,
            "workbook_type": job.workbook_type,
            "status": "skipped",
            "sheet_count": 0,
            "metric_count": 0,
            "notes": f"{type(exc).__name__}: {exc}",
        }
    except Exception as exc:
        connection.execute(
            "UPDATE excel_workbooks SET extraction_status = ?, notes = ? WHERE excel_workbook_id = ?",
            ("error", f"{PARSER_VERSION}; {type(exc).__name__}: {exc}", job.excel_workbook_id),
        )
        return {
            "excel_workbook_id": job.excel_workbook_id,
            "filegroup_id": job.filegroup_id,
            "test_code": job.test_code,
            "workbook_type": job.workbook_type,
            "status": "error",
            "sheet_count": 0,
            "metric_count": 0,
            "notes": f"{type(exc).__name__}: {exc}",
        }


def write_summary(rows: list[dict[str, Any]]) -> None:
    SUMMARY_CSV.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_csv(SUMMARY_CSV, index=False)


def main() -> None:
    args = parse_args()
    warnings.filterwarnings("ignore", message="Cannot parse header or footer so it will be ignored")
    connection = sqlite3.connect(resolve_repo_path(args.db))
    connection.row_factory = sqlite3.Row
    ensure_excel_catalog_schema(connection)
    jobs = load_jobs(connection, args)
    results: list[dict[str, Any]] = []
    started = time.perf_counter()
    commit_interval = max(1, int(args.commit_interval or 1))
    for index, job in enumerate(jobs, start=1):
        results.append(process_job(connection, job))
        if index % commit_interval == 0:
            connection.commit()
    connection.commit()
    ensure_excel_catalog_schema(connection)
    connection.commit()
    write_summary(results)
    summary = {
        "processed": len(results),
        "done": sum(1 for row in results if row["status"] == "done"),
        "skipped": sum(1 for row in results if row["status"] == "skipped"),
        "error": sum(1 for row in results if row["status"] == "error"),
        "sheet_rows": sum(int(row["sheet_count"]) for row in results),
        "metric_rows": sum(int(row["metric_count"]) for row in results),
        "commit_interval": commit_interval,
        "elapsed_seconds": round(time.perf_counter() - started, 2),
        "generated_at": utc_now_iso(),
        "parser_version": PARSER_VERSION,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    connection.close()


if __name__ == "__main__":
    main()
